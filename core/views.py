from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import Group, User
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Q, Sum, Count
from django.views.decorators.cache import cache_control
from django.conf import settings
from django.core.files import File

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from .models import *
from .populate import *
from .forms import *
from .utils import *
from .decorators import unauthenticated_user, allowed_users

import datetime


# Create your views here.

def populate_db(request):
    populate_marque(5)
    return HttpResponse("Test de Données")


def handel404(request, exception):
    return render(request, 'core/404.html')


def handel500(request):
    return render(request, 'core/500.html')


def logout_user(request):
    logout(request)
    return redirect('core:login')


@login_required(login_url='core:login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def home(request):
    return render(request, 'core/home.html')


@unauthenticated_user
def login_page(request):
    # Formulaire validé
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('core:home')
        else:
            messages.warning(request, "Le nom d'utilisateur ou le mot de passe est incorrect.", extra_tags='warning')
            return redirect('core:login')

    return render(request, 'core/login.html')


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def register(request):
    # Active menu
    menu = 'config'
    i_menu = 'register'

    # On recupère tous les utilisateurs
    users = User.objects.all().order_by('username')

    # Formulaire envoyé
    if request.method == 'POST':
        # Recuperation du formulaire
        form = CreateUserForm(request.POST)
        # Si formulaire valide
        if form.is_valid():
            user = form.save()
            group = Group.objects.get(name='staff')
            user.groups.add(group)
            messages.success(request, "L'utilisateur {} a été crée dans la base de données.".format(
                form.cleaned_data['username']),
                             extra_tags='success')
            return redirect('core:login')
        else:
            messages.error(request, "Veuillez remplir tous les champs. Les mots de passe doivent être identiques.",
                           extra_tags='danger')
            return redirect('core:register')

    context = {
        'menu': menu,
        'i_menu': i_menu,
        'users': users
    }
    return render(request, 'core/register.html', context)


@login_required(login_url='core:login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def dashboard(request):
    menu = "dashboard"
    # Choix de l'utilisateur et année
    choice_pi = "Brevets"
    annee_pi = datetime.date.today().year - 1
    type_1 = "Delivrance"
    type_2 = "Demande"

    if request.method == 'POST':
        # recupere le formulaire
        form = DashFilterForm(request.POST)
        # Formulaire valide
        if form.is_valid():
            choice_pi = form.cleaned_data['pi']
            annee_pi = form.cleaned_data['annee']
            # Modification des types
            if choice_pi != 'Brevets':
                type_1 = 'Enregistrement'
        else:
            messages.warning(request, "Veuillez remplir tous les champs du formulaire s'il vous plaît.",
                             extra_tags='warning')
            return redirect('core:dashboard')

    # Vérifie si les 2 années existent dans la BD
    if check_last_year(choice_pi, annee_pi, "recent") & check_last_year(choice_pi, annee_pi, "one"):
        # Totaux
        total_pi_actuel = get_count_pi(choice_pi, "tous", "toutes", annee_pi, "tous")
        total_pi_ancien = get_count_pi(choice_pi, "tous", "toutes", annee_pi - 1, "tous")
        total_pi_etran_actuel = get_count_pi(choice_pi, "tous", "toutes", annee_pi, "Etrangers")
        total_pi_etran_ancien = get_count_pi(choice_pi, "tous", "toutes", annee_pi - 1, "Etrangers")
        total_pi_nat_actuel = get_count_pi(choice_pi, "tous", "toutes", annee_pi, "Nationaux")
        total_pi_nat_ancien = get_count_pi(choice_pi, "tous", "toutes", annee_pi - 1, "Nationaux")
        # Pourcentage
        try:
            pourcentage_total = round(
                (((float(total_pi_actuel) - float(total_pi_ancien)) / float(total_pi_ancien)) * 100),
                2)
            pourcentage_etran = round(
                (((float(total_pi_etran_actuel) - float(total_pi_etran_ancien)) / float(total_pi_etran_ancien)) * 100),
                2)
            pourcentage_nat = round(
                (((float(total_pi_nat_actuel) - float(total_pi_nat_ancien)) / float(total_pi_nat_ancien)) * 100), 2)
        except ZeroDivisionError:
            messages.warning(request,
                             "Certaines données de {0} manquent pour l'année {1}.".format(choice_pi, annee_pi - 1),
                             extra_tags='warning')
            return redirect('core:dashboard')
        print(type_1)
        # Totaux type 1
        total_type_1_actuel = get_count_pi(choice_pi, type_1, "toutes", annee_pi, "tous")
        total_type_1_ancien = get_count_pi(choice_pi, type_1, "toutes", annee_pi - 1, "tous")
        total_type_1_etran_actuel = get_count_pi(choice_pi, type_1, "toutes", annee_pi, "Etrangers")
        total_type_1_etran_ancien = get_count_pi(choice_pi, type_1, "toutes", annee_pi - 1, "Etrangers")
        total_type_1_nat_actuel = get_count_pi(choice_pi, type_1, "toutes", annee_pi, "Nationaux")
        total_type_1_nat_ancien = get_count_pi(choice_pi, type_1, "toutes", annee_pi - 1, "Nationaux")
        # Pourcentage type 1
        try:
            pourcentage_total_type_1 = round(
                (((float(total_type_1_actuel) - float(total_type_1_ancien)) / float(total_type_1_ancien)) * 100), 2)
            pourcentage_etran_type_1 = round(
                (((float(total_type_1_etran_actuel) - float(total_type_1_etran_ancien)) / float(
                    total_type_1_etran_ancien)) * 100), 2)
            pourcentage_nat_type_1 = round(
                (((float(total_type_1_nat_actuel) - float(total_type_1_nat_ancien)) / float(
                    total_type_1_nat_ancien)) * 100), 2)
        except ZeroDivisionError:
            messages.warning(request,
                             "Certaines données de {0} pour le type {1}, manquent pour l'année {2}".format(choice_pi,
                                                                                                           type_1,
                                                                                                           annee_pi - 1),
                             extra_tags='warning')
            return redirect('core:dashboard')

        # Totaux type 2
        total_type_2_actuel = get_count_pi(choice_pi, type_2, "toutes", annee_pi, "tous")
        total_type_2_ancien = get_count_pi(choice_pi, type_2, "toutes", annee_pi - 1, "tous")
        total_type_2_etran_actuel = get_count_pi(choice_pi, type_2, "toutes", annee_pi, "Etrangers")
        total_type_2_etran_ancien = get_count_pi(choice_pi, type_2, "toutes", annee_pi - 1, "Etrangers")
        total_type_2_nat_actuel = get_count_pi(choice_pi, type_2, "toutes", annee_pi, "Nationaux")
        total_type_2_nat_ancien = get_count_pi(choice_pi, type_2, "toutes", annee_pi - 1, "Nationaux")
        # Pourcentage type 2
        try:
            pourcentage_total_type_2 = round(
                (((float(total_type_2_actuel) - float(total_type_2_ancien)) / float(total_type_2_ancien)) * 100), 2)
            pourcentage_etran_type_2 = round(
                (((float(total_type_2_etran_actuel) - float(total_type_2_etran_ancien)) / float(
                    total_type_2_etran_ancien)) * 100), 2)
            pourcentage_nat_type_2 = round(
                (((float(total_type_2_nat_actuel) - float(total_type_2_nat_ancien)) / float(
                    total_type_2_nat_ancien)) * 100), 2)
        except ZeroDivisionError:
            messages.warning(request,
                             "Certaines données de {0} pour le type {1}, manquent pour l'année {2}".format(choice_pi,
                                                                                                           type_2,
                                                                                                           annee_pi - 1),
                             extra_tags='warning')
            return redirect('core:dashboard')

        # Totaux Pays
        total_pays = get_count_countries(choice_pi, "tous", annee_pi, "tous")
        total_pays_etran = get_count_countries(choice_pi, "tous", annee_pi, "Etrangers")
        total_pays_nat = get_count_countries(choice_pi, "tous", annee_pi, "Nationaux")
        # Top 5 pays
        top_5_pays = get_top_countries(choice_pi, "tous", annee_pi, "tous")

        # Graphique
        # Pie Chart
        proportion_type_1 = round(((float(total_type_1_actuel) / float(total_pi_actuel)) * 100), 2)
        proportion_type_2 = round(((float(total_type_2_actuel) / float(total_pi_actuel)) * 100), 2)
        # Line Chart
        if check_last_year(choice_pi, annee_pi, "all"):
            labels = [annee_pi - i for i in range(0, 5)]
            data_total = [
                get_count_pi(choice_pi, "tous", "toutes", annee_pi - i, "tous")
                for i in range(0, 5)
            ]
            data_type_1 = [
                get_count_pi(choice_pi, type_1, "toutes", annee_pi - i, "tous")
                for i in range(0, 5)
            ]
            data_type_2 = [
                get_count_pi(choice_pi, type_2, "toutes", annee_pi - i, "tous")
                for i in range(0, 5)
            ]
        else:
            labels = []
            data_total = []
            data_type_1 = []
            data_type_2 = []

        context = {
            'menu': menu,
            'years': sorted(get_listing_year(), reverse=True),
            'choice_pi': choice_pi,
            'annee_pi': annee_pi,
            'type_1': type_1,
            'type_2': type_2,
            'total_pi_actuel': total_pi_actuel,
            'total_pi_etran_actuel': total_pi_etran_actuel,
            'total_pi_nat_actuel': total_pi_nat_actuel,
            'pourcentage_total': pourcentage_total,
            'pourcentage_etran': pourcentage_etran,
            'pourcentage_nat': pourcentage_nat,
            'total_type_1_actuel': total_type_1_actuel,
            'total_type_1_etran_actuel': total_type_1_etran_actuel,
            'total_type_1_nat_actuel': total_type_1_nat_actuel,
            'pourcentage_total_type_1': pourcentage_total_type_1,
            'pourcentage_etran_type_1': pourcentage_etran_type_1,
            'pourcentage_nat_type_1': pourcentage_nat_type_1,
            'total_type_2_actuel': total_type_2_actuel,
            'total_type_2_etran_actuel': total_type_2_etran_actuel,
            'total_type_2_nat_actuel': total_type_2_nat_actuel,
            'pourcentage_total_type_2': pourcentage_total_type_2,
            'pourcentage_etran_type_2': pourcentage_etran_type_2,
            'pourcentage_nat_type_2': pourcentage_nat_type_2,
            'total_pays': total_pays,
            'total_pays_etran': total_pays_etran,
            'total_pays_nat': total_pays_nat,
            'top_5_pays': top_5_pays,
            'proportion_type_1': proportion_type_1,
            'proportion_type_2': proportion_type_2,
            'labels': sorted(labels),
            'data_total': data_total[::-1],
            'data_type_1': data_type_1[::-1],
            'data_type_2': data_type_2[::-1],
        }
    else:
        context = {
            'menu': menu,
            'years': sorted(get_listing_year(), reverse=True),
            "non_exist": True,
            "choice_pi": choice_pi,
            "annee_pi": annee_pi,
            "annee_passee": annee_pi - 1,
        }
    return render(request, 'core/dashboard.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def details_dashboard(request, choice_pi, type_pi, annee_pi):
    menu = "dashboard"
    # On s'assure que le type est conforme
    if (choice_pi not in ['Brevets', 'DMI', 'Marques']) or (type_pi not in ['Enregistrement', 'Demande', 'Delivrance']):
        messages.error(request, f"Impossible d'obtenir des données pour {choice_pi} et {type_pi}.", extra_tags='danger')
        return redirect('core:dashboard')
    # Choix de la voie
    if choice_pi == "Brevets":
        voie_1 = 'PCT'
        voie_2 = 'CP'
    elif choice_pi == 'DMI':
        voie_1 = 'Etrangers'
        voie_2 = 'Nationaux'
    else:
        voie_1 = 'Classe Produits'
        voie_2 = 'Classe Services'

    # Vérifie si les 2 années existent dans la BD
    if check_last_year(choice_pi, annee_pi, "recent") & check_last_year(choice_pi, annee_pi, "one"):
        # Totaux
        total_type_1_actuel = get_count_pi(choice_pi, type_pi, "toutes", annee_pi, "tous")
        total_type_1_ancien = get_count_pi(choice_pi, type_pi, "toutes", annee_pi - 1, "tous")
        total_type_1_etran_actuel = get_count_pi(choice_pi, type_pi, "toutes", annee_pi, "Etrangers")
        total_type_1_etran_ancien = get_count_pi(choice_pi, type_pi, "toutes", annee_pi - 1, "Etrangers")
        total_type_1_nat_actuel = get_count_pi(choice_pi, type_pi, "toutes", annee_pi, "Nationaux")
        total_type_1_nat_ancien = get_count_pi(choice_pi, type_pi, "toutes", annee_pi - 1, "Nationaux")
        # Pourcentage Totaux
        pourcentage_total_type_1 = round(
            (((float(total_type_1_actuel) - float(total_type_1_ancien)) / float(total_type_1_ancien)) * 100), 2)
        pourcentage_etran_type_1 = round(
            (((float(total_type_1_etran_actuel) - float(total_type_1_etran_ancien)) / float(
                total_type_1_etran_ancien)) * 100), 2)
        pourcentage_nat_type_1 = round(
            (((float(total_type_1_nat_actuel) - float(total_type_1_nat_ancien)) / float(
                total_type_1_nat_ancien)) * 100), 2)

        # Totaux voie 1
        total_voie_1_actuel = get_count_pi(choice_pi, type_pi, voie_1, annee_pi, "tous")
        total_voie_1_ancien = get_count_pi(choice_pi, type_pi, voie_1, annee_pi - 1, "tous")
        total_voie_1_etran_actuel = get_count_pi(choice_pi, type_pi, voie_1, annee_pi, "Etrangers")
        total_voie_1_etran_ancien = get_count_pi(choice_pi, type_pi, voie_1, annee_pi - 1, "Etrangers")
        total_voie_1_nat_actuel = get_count_pi(choice_pi, type_pi, voie_1, annee_pi, "Nationaux")
        total_voie_1_nat_ancien = get_count_pi(choice_pi, type_pi, voie_1, annee_pi - 1, "Nationaux")
        # print(total_voie_1_ancien)
        # Pourcentage voie 1
        if (total_voie_1_actuel != 0) & (total_voie_1_ancien != 0):
            pourcentage_total_voie_1 = round(
                (((float(total_voie_1_actuel) - float(total_voie_1_ancien)) / float(total_voie_1_ancien)) * 100), 2)
        else:
            pourcentage_total_voie_1 = 0
        if (total_voie_1_etran_ancien != 0) & (total_voie_1_etran_actuel != 0):
            pourcentage_etran_voie_1 = round(
                (((float(total_voie_1_etran_actuel) - float(total_voie_1_etran_ancien)) / float(
                    total_voie_1_etran_ancien)) * 100), 2)
        else:
            pourcentage_etran_voie_1 = 0
        if (total_voie_1_nat_ancien != 0) & (total_voie_1_nat_actuel != 0):
            pourcentage_nat_voie_1 = round(
                (((float(total_voie_1_nat_actuel) - float(total_voie_1_nat_ancien)) / float(
                    total_voie_1_nat_ancien)) * 100), 2)
        else:
            pourcentage_nat_voie_1 = 0

        # Totaux voie
        total_voie_2_actuel = get_count_pi(choice_pi, type_pi, voie_2, annee_pi, "tous")
        total_voie_2_ancien = get_count_pi(choice_pi, type_pi, voie_2, annee_pi - 1, "tous")
        total_voie_2_etran_actuel = get_count_pi(choice_pi, type_pi, voie_2, annee_pi, "Etrangers")
        total_voie_2_etran_ancien = get_count_pi(choice_pi, type_pi, voie_2, annee_pi - 1, "Etrangers")
        total_voie_2_nat_actuel = get_count_pi(choice_pi, type_pi, voie_2, annee_pi, "Nationaux")
        total_voie_2_nat_ancien = get_count_pi(choice_pi, type_pi, voie_2, annee_pi - 1, "Nationaux")

        # Pourcentage voie 2
        if (total_voie_2_actuel != 0) & (total_voie_2_ancien != 0):
            pourcentage_total_voie_2 = round(
                (((float(total_voie_2_actuel) - float(total_voie_2_ancien)) / float(total_voie_2_ancien)) * 100), 2)
        else:
            pourcentage_total_voie_2 = 0
        if (total_voie_2_etran_ancien != 0) & (total_voie_2_etran_actuel != 0):
            pourcentage_etran_voie_2 = round(
                (((float(total_voie_2_etran_actuel) - float(total_voie_2_etran_ancien)) / float(
                    total_voie_2_etran_ancien)) * 100), 2)
        else:
            pourcentage_etran_voie_2 = 0
        if (total_voie_2_nat_ancien != 0) & (total_voie_2_nat_actuel != 0):
            pourcentage_nat_voie_2 = round(
                (((float(total_voie_2_nat_actuel) - float(total_voie_2_nat_ancien)) / float(
                    total_voie_2_nat_ancien)) * 100), 2)
        else:
            pourcentage_nat_voie_2 = 0

        # Totaux Pays
        total_pays = get_count_countries(choice_pi, type_pi, annee_pi, "tous")
        total_pays_etran = get_count_countries(choice_pi, type_pi, annee_pi, "Etrangers")
        total_pays_nat = get_count_countries(choice_pi, type_pi, annee_pi, "Nationaux")
        # Top 5 pays
        top_5_pays = get_top_countries(choice_pi, type_pi, annee_pi, "tous")

        # Graphique
        # Pie Chart
        if choice_pi != 'DMI':
            proportion_voie_1 = round(((float(total_voie_1_actuel) / float(total_type_1_actuel)) * 100), 2)
            proportion_voie_2 = round(((float(total_voie_2_actuel) / float(total_type_1_actuel)) * 100), 2)
        else:
            proportion_voie_1 = round(((float(total_voie_1_etran_actuel) / float(total_type_1_actuel)) * 100), 2)
            proportion_voie_2 = round(((float(total_voie_1_nat_actuel) / float(total_type_1_actuel)) * 100), 2)
        # Line Chart
        if check_last_year(choice_pi, annee_pi, "all"):
            labels = [annee_pi - i for i in range(0, 5)]
            data_total = [
                get_count_pi(choice_pi, type_pi, "toutes", annee_pi - i, "tous")
                for i in range(0, 5)
            ]
            if choice_pi != 'DMI':
                data_type_1 = [
                    get_count_pi(choice_pi, type_pi, voie_1, annee_pi - i, "tous")
                    for i in range(0, 5)
                ]
                data_type_2 = [
                    get_count_pi(choice_pi, type_pi, voie_2, annee_pi - i, "tous")
                    for i in range(0, 5)
                ]
            else:
                data_type_1 = [
                    get_count_pi(choice_pi, type_pi, voie_1, annee_pi - i, "Etrangers")
                    for i in range(0, 5)
                ]
                data_type_2 = [
                    get_count_pi(choice_pi, type_pi, voie_2, annee_pi - i, "Nationaux")
                    for i in range(0, 5)
                ]
        else:
            labels = []
            data_total = []
            data_type_1 = []
            data_type_2 = []
        # print(get_top_classe(choice_pi, annee_pi))
        context = {
            'menu': menu,
            'choice_pi': choice_pi,
            'type_pi': type_pi,
            'annee_pi': annee_pi,
            'voie_1': voie_1,
            'voie_2': voie_2,
            'total_pi_actuel': total_type_1_actuel,
            'total_pi_etran_actuel': total_type_1_etran_actuel,
            'total_pi_nat_actuel': total_type_1_nat_actuel,
            'pourcentage_total_type_1': pourcentage_total_type_1,
            'pourcentage_etran_type_1': pourcentage_etran_type_1,
            'pourcentage_nat_type_1': pourcentage_nat_type_1,
            'total_voie_1_actuel': total_voie_1_actuel,
            'total_voie_1_etran_actuel': total_voie_1_etran_actuel,
            'total_voie_1_nat_actuel': total_voie_1_nat_actuel,
            'pourcentage_total_voie_1': pourcentage_total_voie_1,
            'pourcentage_etran_voie_1': pourcentage_etran_voie_1,
            'pourcentage_nat_voie_1': pourcentage_nat_voie_1,
            'total_voie_2_actuel': total_voie_2_actuel,
            'total_voie_2_etran_actuel': total_voie_2_etran_actuel,
            'total_voie_2_nat_actuel': total_voie_2_nat_actuel,
            'pourcentage_total_voie_2': pourcentage_total_voie_2,
            'pourcentage_etran_voie_2': pourcentage_etran_voie_2,
            'pourcentage_nat_voie_2': pourcentage_nat_voie_2,
            'total_pays': total_pays,
            'total_pays_etran': total_pays_etran,
            'total_pays_nat': total_pays_nat,
            'top_5_pays': top_5_pays,
            'proportion_voie_1': proportion_voie_1,
            'proportion_voie_2': proportion_voie_2,
            'labels': sorted(labels),
            'data_total': data_total[::-1],
            'data_type_1': data_type_1[::-1],
            'data_type_2': data_type_2[::-1],
            'classes': get_top_classe(choice_pi, annee_pi)
        }
    else:
        context = {
            'menu': menu,
            'years': sorted(get_listing_year(), reverse=True),
            "non_exist": True,
            "choice_pi": choice_pi,
            "annee_pi": annee_pi,
            "annee_passee": annee_pi - 1,
        }
    return render(request, 'core/details_dashboard.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def export_file_oapi(request):
    # Active menu
    menu = 'config'
    i_menu = 'oapi'
    # path_file = "C:\\Users\\stefn\\OneDrive\\Bureau\\Fichiers OAPI\\"

    # On récupère l'historique
    history = Historique.objects.all().order_by('-date_ajout')

    if request.method == 'POST':
        form = ExportIFileForm(request.POST)
        if form.is_valid():
            # on recupere les données
            pi = form.cleaned_data['pi']
            annee = form.cleaned_data['annee']
            type_pi = form.cleaned_data['type_pi']
            # Données de Brevets
            if pi == 'Brevets':
                # On recupere les données de brevets
                datas = Brevets.objects.filter(Q(type_brevets=type_pi) & Q(annee_brevets=annee)) \
                    .order_by('pays_brevets__nom_pays')
                # S'il y a des données
                if len(datas) > 0:
                    wb = Workbook()
                    # Fichier Global
                    ws = wb.active
                    ws.title = "Global"
                    ws.cell(1, 1, str(type_pi).upper() + " " + str(pi).upper()
                            + " - " + " " + str(annee) + " : " + str(ws.title).upper()).font = Font(bold=True)
                    # Fichier Nationaux
                    ws_nat = wb.create_sheet("Nationaux")
                    ws_nat.cell(1, 1, str(type_pi).upper() + " " + str(pi).upper()
                                + " " + str(annee) + " - " + str(ws_nat.title).upper()).font = Font(bold=True)
                    ws_etran = wb.create_sheet("Etrangers")
                    ws_etran.cell(1, 1, str(type_pi).upper() + " " + str(pi).upper()
                                  + " " + str(annee) + " - " + str(ws_etran.title).upper()).font = Font(bold=True)

                    # Compteur de lignes, colonnes, PCT, CP et TATAUX
                    ligne = 5
                    ligne_nat = 5
                    ligne_etran = 5
                    numero = 1
                    numero_nat = 1
                    numero_etran = 1
                    somme_cp = 0
                    somme_cp_nat = 0
                    somme_cp_etran = 0
                    somme_pct = 0
                    somme_pct_nat = 0
                    somme_pct_etran = 0
                    total = 0
                    total_nat = 0
                    total_etran = 0
                    # Entête de fichiers
                    ws.cell(4, 4, "PCT")
                    ws.cell(4, 5, "CP")
                    ws.cell(4, 6, "TOTAL")
                    # Nationaux
                    ws_nat.cell(4, 4, "PCT")
                    ws_nat.cell(4, 5, "CP")
                    ws_nat.cell(4, 6, "TOTAL")
                    # Etrangers
                    ws_etran.cell(4, 4, "PCT")
                    ws_etran.cell(4, 5, "CP")
                    ws_etran.cell(4, 6, "TOTAL")
                    # Remplissage des cellules
                    for data in datas:
                        # Global
                        ws.cell(ligne, 1, numero)
                        ws.cell(ligne, 2, data.pays_brevets.code_pays)
                        ws.cell(ligne, 3, data.pays_brevets.nom_pays)
                        ws.cell(ligne, 4, data.nombre_pct)
                        ws.cell(ligne, 5, data.nombre_cp)
                        ws.cell(ligne, 6, data.nombre_cp + data.nombre_pct)
                        somme_cp += data.nombre_cp
                        somme_pct += data.nombre_pct
                        total += (data.nombre_pct + data.nombre_cp)
                        ligne += 1
                        numero += 1
                        # Nationaux
                        if data.pays_brevets.specification_pays == 'Nationaux':
                            ws_nat.cell(ligne_nat, 1, numero_nat)
                            ws_nat.cell(ligne_nat, 2, data.pays_brevets.code_pays)
                            ws_nat.cell(ligne_nat, 3, data.pays_brevets.nom_pays)
                            ws_nat.cell(ligne_nat, 4, data.nombre_pct)
                            ws_nat.cell(ligne_nat, 5, data.nombre_cp)
                            ws_nat.cell(ligne_nat, 6, data.nombre_cp + data.nombre_pct)
                            somme_cp_nat += data.nombre_cp
                            somme_pct_nat += data.nombre_pct
                            total_nat += (data.nombre_pct + data.nombre_cp)
                            ligne_nat += 1
                            numero_nat += 1
                        else:
                            # Etrangers
                            ws_etran.cell(ligne_etran, 1, numero_etran)
                            ws_etran.cell(ligne_etran, 2, data.pays_brevets.code_pays)
                            ws_etran.cell(ligne_etran, 3, data.pays_brevets.nom_pays)
                            ws_etran.cell(ligne_etran, 4, data.nombre_pct)
                            ws_etran.cell(ligne_etran, 5, data.nombre_cp)
                            ws_etran.cell(ligne_etran, 6, data.nombre_cp + data.nombre_pct)
                            somme_cp_etran += data.nombre_cp
                            somme_pct_etran += data.nombre_pct
                            total_etran += (data.nombre_pct + data.nombre_cp)
                            ligne_etran += 1
                            numero_etran += 1
                    # Totaux des fichiers
                    # Global
                    ws.cell(ligne, 4, somme_pct)
                    ws.cell(ligne, 5, somme_cp)
                    ws.cell(ligne, 6, total)
                    # Nationaux
                    ws_nat.cell(ligne_nat, 4, somme_pct_nat)
                    ws_nat.cell(ligne_nat, 5, somme_cp_nat)
                    ws_nat.cell(ligne_nat, 6, total_nat)
                    # Etrangers
                    ws_etran.cell(ligne_etran, 4, somme_pct_etran)
                    ws_etran.cell(ligne_etran, 5, somme_cp_etran)
                    ws_etran.cell(ligne_etran, 6, total_etran)

                    # Sauvergarde du fichier
                    wb.save(settings.MEDIA_ROOT + "/" + pi + " - " + type_pi + " " + str(annee) + ".xlsx")
                    file = File(open(settings.MEDIA_ROOT + "/" + pi + " - " + type_pi + " " + str(annee) + ".xlsx", mode='rb'), name=pi + ' - ' + type_pi + ' ' + str(annee) + ".xlsx")
                    messages.success(request, "Les données ont été exportées sous forme de fichier excel.",
                                     extra_tags='success')
                    # On enregistre l'historique
                    h = Historique.objects.create(type_fichier="OAPI", nature_fichier=type_pi, type_pi=pi,
                                                  upload_file=file, annee=annee)
                    file.close()
                    h.save()
                    return redirect('core:export_file_oapi')
                else:
                    messages.warning(request, "Il n'existe aucun {0} de type {1} pour l'année {2} dans la base de"
                                              " données.".format(pi, type_pi, annee), extra_tags='warning')
                    return redirect('core:export_file_oapi')
            # Données de DMI
            if pi == 'DMI':
                # On recupere les données de brevets
                datas = DMI.objects.filter(Q(type_dmi=type_pi) & Q(annee_dmi=annee)) \
                    .order_by('pays_dmi__nom_pays')
                # S'il y a des données
                if len(datas) > 0:
                    wb = Workbook()
                    # Fichier Global
                    ws = wb.active
                    ws.title = "Global"
                    ws.cell(1, 1, str(type_pi).upper() + " " + str(pi).upper()
                            + " - " + " " + str(annee) + " : " + str(ws.title).upper()).font = Font(bold=True)
                    # Fichier Nationaux
                    ws_nat = wb.create_sheet("Nationaux")
                    ws_nat.cell(1, 1, str(type_pi).upper() + " " + str(pi).upper()
                                + " " + str(annee) + " - " + str(ws_nat.title).upper()).font = Font(bold=True)
                    ws_etran = wb.create_sheet("Etrangers")
                    ws_etran.cell(1, 1, str(type_pi).upper() + " " + str(pi).upper()
                                  + " " + str(annee) + " - " + str(ws_etran.title).upper()).font = Font(bold=True)

                    # Compteur de lignes, colonnes, Nbre et Classes
                    ligne = 5
                    ligne_nat = 5
                    ligne_etran = 5
                    numero = 1
                    numero_nat = 1
                    numero_etran = 1
                    somme = 0
                    somme_nat = 0
                    somme_etran = 0
                    total = 0
                    total_nat = 0
                    total_etran = 0
                    # Entête de fichiers
                    ws.cell(4, 4, "Nbre")
                    ws.cell(4, 5, "Classes")
                    ws.cell(4, 11, "1")
                    ws.cell(4, 12, "2")
                    ws.cell(4, 13, "3")
                    ws.cell(4, 14, "4")
                    ws.cell(4, 15, "5")
                    ws.cell(4, 16, "6")
                    ws.cell(4, 17, "7")
                    ws.cell(4, 18, "8")
                    ws.cell(4, 19, "9")
                    ws.cell(4, 20, "10")
                    ws.cell(4, 21, "11")
                    ws.cell(4, 22, "12")
                    ws.cell(4, 23, "13")
                    ws.cell(4, 24, "14")
                    ws.cell(4, 25, "15")
                    ws.cell(4, 26, "16")
                    ws.cell(4, 27, "17")
                    ws.cell(4, 28, "18")
                    ws.cell(4, 29, "19")
                    ws.cell(4, 30, "20")
                    ws.cell(4, 31, "21")
                    ws.cell(4, 32, "22")
                    ws.cell(4, 33, "23")
                    ws.cell(4, 34, "24")
                    ws.cell(4, 35, "25")
                    ws.cell(4, 36, "26")
                    ws.cell(4, 37, "27")
                    ws.cell(4, 38, "28")
                    ws.cell(4, 39, "29")
                    ws.cell(4, 40, "30")
                    ws.cell(4, 41, "31")
                    ws.cell(4, 42, "32")
                    ws.cell(4, 43, "Total")
                    # Nationaux
                    ws_nat.cell(4, 4, "Nbre")
                    ws_nat.cell(4, 5, "Classes")
                    ws_nat.cell(4, 11, "1")
                    ws_nat.cell(4, 12, "2")
                    ws_nat.cell(4, 13, "3")
                    ws_nat.cell(4, 14, "4")
                    ws_nat.cell(4, 15, "5")
                    ws_nat.cell(4, 16, "6")
                    ws_nat.cell(4, 17, "7")
                    ws_nat.cell(4, 18, "8")
                    ws_nat.cell(4, 19, "9")
                    ws_nat.cell(4, 20, "10")
                    ws_nat.cell(4, 21, "11")
                    ws_nat.cell(4, 22, "12")
                    ws_nat.cell(4, 23, "13")
                    ws_nat.cell(4, 24, "14")
                    ws_nat.cell(4, 25, "15")
                    ws_nat.cell(4, 26, "16")
                    ws_nat.cell(4, 27, "17")
                    ws_nat.cell(4, 28, "18")
                    ws_nat.cell(4, 29, "19")
                    ws_nat.cell(4, 30, "20")
                    ws_nat.cell(4, 31, "21")
                    ws_nat.cell(4, 32, "22")
                    ws_nat.cell(4, 33, "23")
                    ws_nat.cell(4, 34, "24")
                    ws_nat.cell(4, 35, "25")
                    ws_nat.cell(4, 36, "26")
                    ws_nat.cell(4, 37, "27")
                    ws_nat.cell(4, 38, "28")
                    ws_nat.cell(4, 39, "29")
                    ws_nat.cell(4, 40, "30")
                    ws_nat.cell(4, 41, "31")
                    ws_nat.cell(4, 42, "32")
                    ws_nat.cell(4, 43, "Total")
                    # Etrangers
                    ws_etran.cell(4, 4, "Nbre")
                    ws_etran.cell(4, 5, "Classes")
                    ws_etran.cell(4, 11, "1")
                    ws_etran.cell(4, 12, "2")
                    ws_etran.cell(4, 13, "3")
                    ws_etran.cell(4, 14, "4")
                    ws_etran.cell(4, 15, "5")
                    ws_etran.cell(4, 16, "6")
                    ws_etran.cell(4, 17, "7")
                    ws_etran.cell(4, 18, "8")
                    ws_etran.cell(4, 19, "9")
                    ws_etran.cell(4, 20, "10")
                    ws_etran.cell(4, 21, "11")
                    ws_etran.cell(4, 22, "12")
                    ws_etran.cell(4, 23, "13")
                    ws_etran.cell(4, 24, "14")
                    ws_etran.cell(4, 25, "15")
                    ws_etran.cell(4, 26, "16")
                    ws_etran.cell(4, 27, "17")
                    ws_etran.cell(4, 28, "18")
                    ws_etran.cell(4, 29, "19")
                    ws_etran.cell(4, 30, "20")
                    ws_etran.cell(4, 31, "21")
                    ws_etran.cell(4, 32, "22")
                    ws_etran.cell(4, 33, "23")
                    ws_etran.cell(4, 34, "24")
                    ws_etran.cell(4, 35, "25")
                    ws_etran.cell(4, 36, "26")
                    ws_etran.cell(4, 37, "27")
                    ws_etran.cell(4, 38, "28")
                    ws_etran.cell(4, 39, "29")
                    ws_etran.cell(4, 40, "30")
                    ws_etran.cell(4, 41, "31")
                    ws_etran.cell(4, 42, "32")
                    ws_etran.cell(4, 43, "Total")
                    # Remplissage des cellules
                    for data in datas:
                        # Global
                        ws.cell(ligne, 1, numero)
                        ws.cell(ligne, 2, data.pays_dmi.code_pays)
                        ws.cell(ligne, 3, data.pays_dmi.nom_pays)
                        ws.cell(ligne, 4, data.nombre_dmi)
                        ws.cell(ligne, 5, data.somme_classes)

                        ws.cell(ligne, 8, numero)
                        ws.cell(ligne, 9, data.pays_dmi.code_pays)
                        ws.cell(ligne, 10, data.pays_dmi.nom_pays)
                        ws.cell(ligne, 11, data.classe_1)
                        ws.cell(ligne, 12, data.classe_2)
                        ws.cell(ligne, 13, data.classe_3)
                        ws.cell(ligne, 14, data.classe_4)
                        ws.cell(ligne, 15, data.classe_5)
                        ws.cell(ligne, 16, data.classe_6)
                        ws.cell(ligne, 17, data.classe_7)
                        ws.cell(ligne, 18, data.classe_8)
                        ws.cell(ligne, 19, data.classe_9)
                        ws.cell(ligne, 20, data.classe_10)
                        ws.cell(ligne, 21, data.classe_11)
                        ws.cell(ligne, 22, data.classe_12)
                        ws.cell(ligne, 23, data.classe_13)
                        ws.cell(ligne, 24, data.classe_14)
                        ws.cell(ligne, 25, data.classe_15)
                        ws.cell(ligne, 26, data.classe_16)
                        ws.cell(ligne, 27, data.classe_17)
                        ws.cell(ligne, 28, data.classe_18)
                        ws.cell(ligne, 29, data.classe_19)
                        ws.cell(ligne, 30, data.classe_20)
                        ws.cell(ligne, 31, data.classe_21)
                        ws.cell(ligne, 32, data.classe_22)
                        ws.cell(ligne, 33, data.classe_23)
                        ws.cell(ligne, 34, data.classe_24)
                        ws.cell(ligne, 35, data.classe_25)
                        ws.cell(ligne, 36, data.classe_26)
                        ws.cell(ligne, 37, data.classe_27)
                        ws.cell(ligne, 38, data.classe_28)
                        ws.cell(ligne, 39, data.classe_29)
                        ws.cell(ligne, 40, data.classe_30)
                        ws.cell(ligne, 41, data.classe_31)
                        ws.cell(ligne, 42, data.classe_32)
                        ws.cell(ligne, 43, data.somme_classes)
                        ligne += 1
                        numero += 1
                        somme += data.nombre_dmi
                        total += data.somme_classes
                        # Nationaux
                        if data.pays_dmi.specification_pays == 'Nationaux':
                            ws_nat.cell(ligne_nat, 1, numero_nat)
                            ws_nat.cell(ligne_nat, 2, data.pays_dmi.code_pays)
                            ws_nat.cell(ligne_nat, 3, data.pays_dmi.nom_pays)
                            ws_nat.cell(ligne_nat, 4, data.nombre_dmi)
                            ws_nat.cell(ligne_nat, 5, data.somme_classes)

                            ws_nat.cell(ligne_nat, 8, numero_nat)
                            ws_nat.cell(ligne_nat, 9, data.pays_dmi.code_pays)
                            ws_nat.cell(ligne_nat, 10, data.pays_dmi.nom_pays)
                            ws_nat.cell(ligne_nat, 11, data.classe_1)
                            ws_nat.cell(ligne_nat, 12, data.classe_2)
                            ws_nat.cell(ligne_nat, 13, data.classe_3)
                            ws_nat.cell(ligne_nat, 14, data.classe_4)
                            ws_nat.cell(ligne_nat, 15, data.classe_5)
                            ws_nat.cell(ligne_nat, 16, data.classe_6)
                            ws_nat.cell(ligne_nat, 17, data.classe_7)
                            ws_nat.cell(ligne_nat, 18, data.classe_8)
                            ws_nat.cell(ligne_nat, 19, data.classe_9)
                            ws_nat.cell(ligne_nat, 20, data.classe_10)
                            ws_nat.cell(ligne_nat, 21, data.classe_11)
                            ws_nat.cell(ligne_nat, 22, data.classe_12)
                            ws_nat.cell(ligne_nat, 23, data.classe_13)
                            ws_nat.cell(ligne_nat, 24, data.classe_14)
                            ws_nat.cell(ligne_nat, 25, data.classe_15)
                            ws_nat.cell(ligne_nat, 26, data.classe_16)
                            ws_nat.cell(ligne_nat, 27, data.classe_17)
                            ws_nat.cell(ligne_nat, 28, data.classe_18)
                            ws_nat.cell(ligne_nat, 29, data.classe_19)
                            ws_nat.cell(ligne_nat, 30, data.classe_20)
                            ws_nat.cell(ligne_nat, 31, data.classe_21)
                            ws_nat.cell(ligne_nat, 32, data.classe_22)
                            ws_nat.cell(ligne_nat, 33, data.classe_23)
                            ws_nat.cell(ligne_nat, 34, data.classe_24)
                            ws_nat.cell(ligne_nat, 35, data.classe_25)
                            ws_nat.cell(ligne_nat, 36, data.classe_26)
                            ws_nat.cell(ligne_nat, 37, data.classe_27)
                            ws_nat.cell(ligne_nat, 38, data.classe_28)
                            ws_nat.cell(ligne_nat, 39, data.classe_29)
                            ws_nat.cell(ligne_nat, 40, data.classe_30)
                            ws_nat.cell(ligne_nat, 41, data.classe_31)
                            ws_nat.cell(ligne_nat, 42, data.classe_32)
                            ws_nat.cell(ligne_nat, 43, data.somme_classes)
                            ligne_nat += 1
                            numero_nat += 1
                            somme_nat += data.nombre_dmi
                            total_nat += data.somme_classes
                        else:
                            # Etrangers
                            ws_etran.cell(ligne_etran, 1, numero_etran)
                            ws_etran.cell(ligne_etran, 2, data.pays_dmi.code_pays)
                            ws_etran.cell(ligne_etran, 3, data.pays_dmi.nom_pays)
                            ws_etran.cell(ligne_etran, 4, data.nombre_dmi)
                            ws_etran.cell(ligne_etran, 5, data.somme_classes)

                            ws_etran.cell(ligne_etran, 8, numero_nat)
                            ws_etran.cell(ligne_etran, 9, data.pays_dmi.code_pays)
                            ws_etran.cell(ligne_etran, 10, data.pays_dmi.nom_pays)
                            ws_etran.cell(ligne_etran, 11, data.classe_1)
                            ws_etran.cell(ligne_etran, 12, data.classe_2)
                            ws_etran.cell(ligne_etran, 13, data.classe_3)
                            ws_etran.cell(ligne_etran, 14, data.classe_4)
                            ws_etran.cell(ligne_etran, 15, data.classe_5)
                            ws_etran.cell(ligne_etran, 16, data.classe_6)
                            ws_etran.cell(ligne_etran, 17, data.classe_7)
                            ws_etran.cell(ligne_etran, 18, data.classe_8)
                            ws_etran.cell(ligne_etran, 19, data.classe_9)
                            ws_etran.cell(ligne_etran, 20, data.classe_10)
                            ws_etran.cell(ligne_etran, 21, data.classe_11)
                            ws_etran.cell(ligne_etran, 22, data.classe_12)
                            ws_etran.cell(ligne_etran, 23, data.classe_13)
                            ws_etran.cell(ligne_etran, 24, data.classe_14)
                            ws_etran.cell(ligne_etran, 25, data.classe_15)
                            ws_etran.cell(ligne_etran, 26, data.classe_16)
                            ws_etran.cell(ligne_etran, 27, data.classe_17)
                            ws_etran.cell(ligne_etran, 28, data.classe_18)
                            ws_etran.cell(ligne_etran, 29, data.classe_19)
                            ws_etran.cell(ligne_etran, 30, data.classe_20)
                            ws_etran.cell(ligne_etran, 31, data.classe_21)
                            ws_etran.cell(ligne_etran, 32, data.classe_22)
                            ws_etran.cell(ligne_etran, 33, data.classe_23)
                            ws_etran.cell(ligne_etran, 34, data.classe_24)
                            ws_etran.cell(ligne_etran, 35, data.classe_25)
                            ws_etran.cell(ligne_etran, 36, data.classe_26)
                            ws_etran.cell(ligne_etran, 37, data.classe_27)
                            ws_etran.cell(ligne_etran, 38, data.classe_28)
                            ws_etran.cell(ligne_etran, 39, data.classe_29)
                            ws_etran.cell(ligne_etran, 40, data.classe_30)
                            ws_etran.cell(ligne_etran, 41, data.classe_31)
                            ws_etran.cell(ligne_etran, 42, data.classe_32)
                            ws_etran.cell(ligne_etran, 43, data.somme_classes)
                            ligne_etran += 1
                            numero_etran += 1
                            somme_etran += data.nombre_dmi
                            total_etran += data.somme_classes
                    # Totaux des fichiers
                    # Global
                    ws.cell(ligne, 4, somme)
                    ws.cell(ligne, 5, total)
                    # Nationaux
                    ws_nat.cell(ligne_nat, 4, somme_nat)
                    ws_nat.cell(ligne_nat, 5, total_nat)
                    # Etrangers
                    ws_etran.cell(ligne_etran, 4, somme_etran)
                    ws_etran.cell(ligne_etran, 5, total_etran)

                    # Sauvergarde du fichier
                    wb.save(settings.MEDIA_ROOT + "/" + pi + " - " + type_pi + " " + str(annee) + ".xlsx")
                    file = File(
                        open(settings.MEDIA_ROOT + "/" + pi + " - " + type_pi + " " + str(annee) + ".xlsx", mode='rb'),
                        name=pi + ' - ' + type_pi + ' ' + str(annee) + ".xlsx")
                    messages.success(request, "Les données ont été exportées sous forme de fichier excel.",
                                     extra_tags='success')
                    # On enregistre l'historique
                    h = Historique.objects.create(type_fichier="OAPI", nature_fichier=type_pi, type_pi=pi,
                                                  upload_file=file, annee=annee)
                    h.save()
                    return redirect('core:export_file_oapi')
                else:
                    messages.warning(request, "Il n'existe aucun {0} de type {1} pour l'année {2} dans la base de"
                                              " données.".format(pi, type_pi, annee), extra_tags='warning')
                    return redirect('core:export_file_oapi')
            # Données Marques
            if pi == 'Marque':
                # On recupere les données de marques
                if type_pi == 'Demande':
                    datas = Marques.objects.filter(Q(type_marques__contains=type_pi) & Q(annee_marques=annee)) \
                        .order_by('pays_marques__nom_pays')
                else:
                    datas = Marques.objects.filter(Q(type_marques__contains='Enreg') & Q(annee_marques=annee)) \
                        .order_by('pays_marques__nom_pays')
                # S'il y a des données
                if len(datas) > 0:
                    # Création du classeur et des feuilles
                    wb = Workbook()
                    if type_pi == 'Demande':
                        sheet_name = type_pi
                    else:
                        sheet_name = 'Enreg'
                    ws = wb.active
                    ws.title = sheet_name + " - Global"
                    ws.cell(1, 1,
                            str(type_pi).upper() + " " + str(pi).upper() + " GLOBAL - " + str(annee).upper()).font = \
                        Font(bold=True)
                    ws_etran_global = wb.create_sheet(sheet_name + " Global - Etranger")
                    ws_etran_global.cell(1, 1, str(type_pi).upper() + " " + str(pi).upper() + " GLOBAL - ETRANGER "
                                         + str(annee).upper()).font = Font(bold=True)
                    ws_reg_etran = wb.create_sheet(sheet_name + " Regionale - Etranger")
                    ws_reg_etran.cell(1, 1, str(type_pi).upper() + " " + str(pi).upper() + " REGIONALE - ETRANGER "
                                      + str(annee).upper()).font = Font(bold=True)
                    ws_reg_nat = wb.create_sheet(sheet_name + " Regionale - Nationale")
                    ws_reg_nat.cell(1, 1, str(type_pi).upper() + " " + str(pi).upper() + " REGIONALE - NATIONALE "
                                    + str(annee).upper()).font = Font(bold=True)
                    ws_reg = wb.create_sheet(sheet_name + " - Voie Regionale")
                    ws_reg.cell(1, 1, str(type_pi).upper() + " " + str(pi).upper() + " - VOIE REGIONALE "
                                + str(annee).upper()).font = Font(bold=True)
                    ws_madrid = wb.create_sheet(sheet_name + " Madrid")
                    ws_madrid.cell(1, 1, str(type_pi).upper() + " " + str(pi).upper() + " MADRID "
                                   + str(annee).upper()).font = Font(bold=True)
                    # Compteur de lignes, colonnes, Nbre et Classes
                    ligne_reg = 5
                    ligne_nat = 5
                    ligne_etran = 5
                    ligne_madrid = 5
                    ligne_etr_glob = 5
                    ligne = 5
                    numero_reg = 1
                    numero_nat = 1
                    numero_etran = 1
                    numero_madrid = 1
                    numero_etr_glob = 1
                    numero = 1
                    somme_reg = 0
                    somme_nat = 0
                    somme_etran = 0
                    somme_madrid = 0
                    somme_etr_glob = 0
                    somme = 0
                    total_reg = 0
                    total_nat = 0
                    total_etran = 0
                    total_madrid = 0
                    total_etr_glob = 0
                    total = 0
                    # Entête de fichiers
                    ws.cell(4, 4, "Nbre")
                    ws.cell(4, 5, "Classes")
                    ws.cell(4, 11, "1")
                    ws.cell(4, 12, "2")
                    ws.cell(4, 13, "3")
                    ws.cell(4, 14, "4")
                    ws.cell(4, 15, "5")
                    ws.cell(4, 16, "6")
                    ws.cell(4, 17, "7")
                    ws.cell(4, 18, "8")
                    ws.cell(4, 19, "9")
                    ws.cell(4, 20, "10")
                    ws.cell(4, 21, "11")
                    ws.cell(4, 22, "12")
                    ws.cell(4, 23, "13")
                    ws.cell(4, 24, "14")
                    ws.cell(4, 25, "15")
                    ws.cell(4, 26, "16")
                    ws.cell(4, 27, "17")
                    ws.cell(4, 28, "18")
                    ws.cell(4, 29, "19")
                    ws.cell(4, 30, "20")
                    ws.cell(4, 31, "21")
                    ws.cell(4, 32, "22")
                    ws.cell(4, 33, "23")
                    ws.cell(4, 34, "24")
                    ws.cell(4, 35, "25")
                    ws.cell(4, 36, "26")
                    ws.cell(4, 37, "27")
                    ws.cell(4, 38, "28")
                    ws.cell(4, 39, "29")
                    ws.cell(4, 40, "30")
                    ws.cell(4, 41, "31")
                    ws.cell(4, 42, "32")
                    ws.cell(4, 43, "33")
                    ws.cell(4, 44, "34")
                    ws.cell(4, 45, "35")
                    ws.cell(4, 46, "36")
                    ws.cell(4, 47, "37")
                    ws.cell(4, 48, "38")
                    ws.cell(4, 49, "39")
                    ws.cell(4, 50, "40")
                    ws.cell(4, 51, "41")
                    ws.cell(4, 52, "42")
                    ws.cell(4, 53, "43")
                    ws.cell(4, 54, "44")
                    ws.cell(4, 55, "45")
                    ws.cell(4, 56, "Total")
                    # Global etrangers
                    ws_etran_global.cell(4, 4, "Nbre")
                    ws_etran_global.cell(4, 5, "Classes")
                    ws_etran_global.cell(4, 11, "1")
                    ws_etran_global.cell(4, 12, "2")
                    ws_etran_global.cell(4, 13, "3")
                    ws_etran_global.cell(4, 14, "4")
                    ws_etran_global.cell(4, 15, "5")
                    ws_etran_global.cell(4, 16, "6")
                    ws_etran_global.cell(4, 17, "7")
                    ws_etran_global.cell(4, 18, "8")
                    ws_etran_global.cell(4, 19, "9")
                    ws_etran_global.cell(4, 20, "10")
                    ws_etran_global.cell(4, 21, "11")
                    ws_etran_global.cell(4, 22, "12")
                    ws_etran_global.cell(4, 23, "13")
                    ws_etran_global.cell(4, 24, "14")
                    ws_etran_global.cell(4, 25, "15")
                    ws_etran_global.cell(4, 26, "16")
                    ws_etran_global.cell(4, 27, "17")
                    ws_etran_global.cell(4, 28, "18")
                    ws_etran_global.cell(4, 29, "19")
                    ws_etran_global.cell(4, 30, "20")
                    ws_etran_global.cell(4, 31, "21")
                    ws_etran_global.cell(4, 32, "22")
                    ws_etran_global.cell(4, 33, "23")
                    ws_etran_global.cell(4, 34, "24")
                    ws_etran_global.cell(4, 35, "25")
                    ws_etran_global.cell(4, 36, "26")
                    ws_etran_global.cell(4, 37, "27")
                    ws_etran_global.cell(4, 38, "28")
                    ws_etran_global.cell(4, 39, "29")
                    ws_etran_global.cell(4, 40, "30")
                    ws_etran_global.cell(4, 41, "31")
                    ws_etran_global.cell(4, 42, "32")
                    ws_etran_global.cell(4, 43, "33")
                    ws_etran_global.cell(4, 44, "34")
                    ws_etran_global.cell(4, 45, "35")
                    ws_etran_global.cell(4, 46, "36")
                    ws_etran_global.cell(4, 47, "37")
                    ws_etran_global.cell(4, 48, "38")
                    ws_etran_global.cell(4, 49, "39")
                    ws_etran_global.cell(4, 50, "40")
                    ws_etran_global.cell(4, 51, "41")
                    ws_etran_global.cell(4, 52, "42")
                    ws_etran_global.cell(4, 53, "43")
                    ws_etran_global.cell(4, 54, "44")
                    ws_etran_global.cell(4, 55, "45")
                    ws_etran_global.cell(4, 56, "Total")
                    # Voie regionale
                    ws_reg.cell(4, 4, "Nbre")
                    ws_reg.cell(4, 5, "Classes")
                    ws_reg.cell(4, 11, "1")
                    ws_reg.cell(4, 12, "2")
                    ws_reg.cell(4, 13, "3")
                    ws_reg.cell(4, 14, "4")
                    ws_reg.cell(4, 15, "5")
                    ws_reg.cell(4, 16, "6")
                    ws_reg.cell(4, 17, "7")
                    ws_reg.cell(4, 18, "8")
                    ws_reg.cell(4, 19, "9")
                    ws_reg.cell(4, 20, "10")
                    ws_reg.cell(4, 21, "11")
                    ws_reg.cell(4, 22, "12")
                    ws_reg.cell(4, 23, "13")
                    ws_reg.cell(4, 24, "14")
                    ws_reg.cell(4, 25, "15")
                    ws_reg.cell(4, 26, "16")
                    ws_reg.cell(4, 27, "17")
                    ws_reg.cell(4, 28, "18")
                    ws_reg.cell(4, 29, "19")
                    ws_reg.cell(4, 30, "20")
                    ws_reg.cell(4, 31, "21")
                    ws_reg.cell(4, 32, "22")
                    ws_reg.cell(4, 33, "23")
                    ws_reg.cell(4, 34, "24")
                    ws_reg.cell(4, 35, "25")
                    ws_reg.cell(4, 36, "26")
                    ws_reg.cell(4, 37, "27")
                    ws_reg.cell(4, 38, "28")
                    ws_reg.cell(4, 39, "29")
                    ws_reg.cell(4, 40, "30")
                    ws_reg.cell(4, 41, "31")
                    ws_reg.cell(4, 42, "32")
                    ws_reg.cell(4, 43, "33")
                    ws_reg.cell(4, 44, "34")
                    ws_reg.cell(4, 45, "35")
                    ws_reg.cell(4, 46, "36")
                    ws_reg.cell(4, 47, "37")
                    ws_reg.cell(4, 48, "38")
                    ws_reg.cell(4, 49, "39")
                    ws_reg.cell(4, 50, "40")
                    ws_reg.cell(4, 51, "41")
                    ws_reg.cell(4, 52, "42")
                    ws_reg.cell(4, 53, "43")
                    ws_reg.cell(4, 54, "44")
                    ws_reg.cell(4, 55, "45")
                    ws_reg.cell(4, 56, "Total")
                    # Etranger
                    ws_reg_etran.cell(4, 4, "Nbre")
                    ws_reg_etran.cell(4, 5, "Classes")
                    ws_reg_etran.cell(4, 11, "1")
                    ws_reg_etran.cell(4, 12, "2")
                    ws_reg_etran.cell(4, 13, "3")
                    ws_reg_etran.cell(4, 14, "4")
                    ws_reg_etran.cell(4, 15, "5")
                    ws_reg_etran.cell(4, 16, "6")
                    ws_reg_etran.cell(4, 17, "7")
                    ws_reg_etran.cell(4, 18, "8")
                    ws_reg_etran.cell(4, 19, "9")
                    ws_reg_etran.cell(4, 20, "10")
                    ws_reg_etran.cell(4, 21, "11")
                    ws_reg_etran.cell(4, 22, "12")
                    ws_reg_etran.cell(4, 23, "13")
                    ws_reg_etran.cell(4, 24, "14")
                    ws_reg_etran.cell(4, 25, "15")
                    ws_reg_etran.cell(4, 26, "16")
                    ws_reg_etran.cell(4, 27, "17")
                    ws_reg_etran.cell(4, 28, "18")
                    ws_reg_etran.cell(4, 29, "19")
                    ws_reg_etran.cell(4, 30, "20")
                    ws_reg_etran.cell(4, 31, "21")
                    ws_reg_etran.cell(4, 32, "22")
                    ws_reg_etran.cell(4, 33, "23")
                    ws_reg_etran.cell(4, 34, "24")
                    ws_reg_etran.cell(4, 35, "25")
                    ws_reg_etran.cell(4, 36, "26")
                    ws_reg_etran.cell(4, 37, "27")
                    ws_reg_etran.cell(4, 38, "28")
                    ws_reg_etran.cell(4, 39, "29")
                    ws_reg_etran.cell(4, 40, "30")
                    ws_reg_etran.cell(4, 41, "31")
                    ws_reg_etran.cell(4, 42, "32")
                    ws_reg_etran.cell(4, 43, "33")
                    ws_reg_etran.cell(4, 44, "34")
                    ws_reg_etran.cell(4, 45, "35")
                    ws_reg_etran.cell(4, 46, "36")
                    ws_reg_etran.cell(4, 47, "37")
                    ws_reg_etran.cell(4, 48, "38")
                    ws_reg_etran.cell(4, 49, "39")
                    ws_reg_etran.cell(4, 50, "40")
                    ws_reg_etran.cell(4, 51, "41")
                    ws_reg_etran.cell(4, 52, "42")
                    ws_reg_etran.cell(4, 53, "43")
                    ws_reg_etran.cell(4, 54, "44")
                    ws_reg_nat.cell(4, 55, "45")
                    ws_reg_etran.cell(4, 56, "Total")
                    # Nationale
                    ws_reg_nat.cell(4, 4, "Nbre")
                    ws_reg_nat.cell(4, 5, "Classes")
                    ws_reg_nat.cell(4, 11, "1")
                    ws_reg_nat.cell(4, 12, "2")
                    ws_reg_nat.cell(4, 13, "3")
                    ws_reg_nat.cell(4, 14, "4")
                    ws_reg_nat.cell(4, 15, "5")
                    ws_reg_nat.cell(4, 16, "6")
                    ws_reg_nat.cell(4, 17, "7")
                    ws_reg_nat.cell(4, 18, "8")
                    ws_reg_nat.cell(4, 19, "9")
                    ws_reg_nat.cell(4, 20, "10")
                    ws_reg_nat.cell(4, 21, "11")
                    ws_reg_nat.cell(4, 22, "12")
                    ws_reg_nat.cell(4, 23, "13")
                    ws_reg_nat.cell(4, 24, "14")
                    ws_reg_nat.cell(4, 25, "15")
                    ws_reg_nat.cell(4, 26, "16")
                    ws_reg_nat.cell(4, 27, "17")
                    ws_reg_nat.cell(4, 28, "18")
                    ws_reg_nat.cell(4, 29, "19")
                    ws_reg_nat.cell(4, 30, "20")
                    ws_reg_nat.cell(4, 31, "21")
                    ws_reg_nat.cell(4, 32, "22")
                    ws_reg_nat.cell(4, 33, "23")
                    ws_reg_nat.cell(4, 34, "24")
                    ws_reg_nat.cell(4, 35, "25")
                    ws_reg_nat.cell(4, 36, "26")
                    ws_reg_nat.cell(4, 37, "27")
                    ws_reg_nat.cell(4, 38, "28")
                    ws_reg_nat.cell(4, 39, "29")
                    ws_reg_nat.cell(4, 40, "30")
                    ws_reg_nat.cell(4, 41, "31")
                    ws_reg_nat.cell(4, 42, "32")
                    ws_reg_nat.cell(4, 43, "33")
                    ws_reg_nat.cell(4, 44, "34")
                    ws_reg_nat.cell(4, 45, "35")
                    ws_reg_nat.cell(4, 46, "36")
                    ws_reg_nat.cell(4, 47, "37")
                    ws_reg_nat.cell(4, 48, "38")
                    ws_reg_nat.cell(4, 49, "39")
                    ws_reg_nat.cell(4, 50, "40")
                    ws_reg_nat.cell(4, 51, "41")
                    ws_reg_nat.cell(4, 52, "42")
                    ws_reg_nat.cell(4, 53, "43")
                    ws_reg_nat.cell(4, 54, "44")
                    ws_reg_nat.cell(4, 55, "45")
                    ws_reg_nat.cell(4, 56, "Total")
                    # Madrid
                    ws_madrid.cell(4, 4, "Nbre")
                    ws_madrid.cell(4, 5, "Classes")
                    ws_madrid.cell(4, 11, "1")
                    ws_madrid.cell(4, 12, "2")
                    ws_madrid.cell(4, 13, "3")
                    ws_madrid.cell(4, 14, "4")
                    ws_madrid.cell(4, 15, "5")
                    ws_madrid.cell(4, 16, "6")
                    ws_madrid.cell(4, 17, "7")
                    ws_madrid.cell(4, 18, "8")
                    ws_madrid.cell(4, 19, "9")
                    ws_madrid.cell(4, 20, "10")
                    ws_madrid.cell(4, 21, "11")
                    ws_madrid.cell(4, 22, "12")
                    ws_madrid.cell(4, 23, "13")
                    ws_madrid.cell(4, 24, "14")
                    ws_madrid.cell(4, 25, "15")
                    ws_madrid.cell(4, 26, "16")
                    ws_madrid.cell(4, 27, "17")
                    ws_madrid.cell(4, 28, "18")
                    ws_madrid.cell(4, 29, "19")
                    ws_madrid.cell(4, 30, "20")
                    ws_madrid.cell(4, 31, "21")
                    ws_madrid.cell(4, 32, "22")
                    ws_madrid.cell(4, 33, "23")
                    ws_madrid.cell(4, 34, "24")
                    ws_madrid.cell(4, 35, "25")
                    ws_madrid.cell(4, 36, "26")
                    ws_madrid.cell(4, 37, "27")
                    ws_madrid.cell(4, 38, "28")
                    ws_madrid.cell(4, 39, "29")
                    ws_madrid.cell(4, 40, "30")
                    ws_madrid.cell(4, 41, "31")
                    ws_madrid.cell(4, 42, "32")
                    ws_madrid.cell(4, 43, "33")
                    ws_madrid.cell(4, 44, "34")
                    ws_madrid.cell(4, 45, "35")
                    ws_madrid.cell(4, 46, "36")
                    ws_madrid.cell(4, 47, "37")
                    ws_madrid.cell(4, 48, "38")
                    ws_madrid.cell(4, 49, "39")
                    ws_madrid.cell(4, 50, "40")
                    ws_madrid.cell(4, 51, "41")
                    ws_madrid.cell(4, 52, "42")
                    ws_madrid.cell(4, 53, "43")
                    ws_madrid.cell(4, 54, "44")
                    ws_madrid.cell(4, 55, "45")
                    ws_madrid.cell(4, 56, "Total")
                    # Remplissage de cellules
                    for data in datas:
                        # Voie régionale
                        if "Regionale" in data.type_marques:
                            # Global
                            ws_reg.cell(ligne_reg, 1, numero_reg)
                            ws_reg.cell(ligne_reg, 2, data.pays_marques.code_pays)
                            ws_reg.cell(ligne_reg, 3, data.pays_marques.nom_pays)
                            ws_reg.cell(ligne_reg, 4, data.nombre_marques)
                            ws_reg.cell(ligne_reg, 5, data.somme_classes)

                            ws_reg.cell(ligne_reg, 8, numero_reg)
                            ws_reg.cell(ligne_reg, 9, data.pays_marques.code_pays)
                            ws_reg.cell(ligne_reg, 10, data.pays_marques.nom_pays)
                            ws_reg.cell(ligne_reg, 11, data.classe_1)
                            ws_reg.cell(ligne_reg, 12, data.classe_2)
                            ws_reg.cell(ligne_reg, 13, data.classe_3)
                            ws_reg.cell(ligne_reg, 14, data.classe_4)
                            ws_reg.cell(ligne_reg, 15, data.classe_5)
                            ws_reg.cell(ligne_reg, 16, data.classe_6)
                            ws_reg.cell(ligne_reg, 17, data.classe_7)
                            ws_reg.cell(ligne_reg, 18, data.classe_8)
                            ws_reg.cell(ligne_reg, 19, data.classe_9)
                            ws_reg.cell(ligne_reg, 20, data.classe_10)
                            ws_reg.cell(ligne_reg, 21, data.classe_11)
                            ws_reg.cell(ligne_reg, 22, data.classe_12)
                            ws_reg.cell(ligne_reg, 23, data.classe_13)
                            ws_reg.cell(ligne_reg, 24, data.classe_14)
                            ws_reg.cell(ligne_reg, 25, data.classe_15)
                            ws_reg.cell(ligne_reg, 26, data.classe_16)
                            ws_reg.cell(ligne_reg, 27, data.classe_17)
                            ws_reg.cell(ligne_reg, 28, data.classe_18)
                            ws_reg.cell(ligne_reg, 29, data.classe_19)
                            ws_reg.cell(ligne_reg, 30, data.classe_20)
                            ws_reg.cell(ligne_reg, 31, data.classe_21)
                            ws_reg.cell(ligne_reg, 32, data.classe_22)
                            ws_reg.cell(ligne_reg, 33, data.classe_23)
                            ws_reg.cell(ligne_reg, 34, data.classe_24)
                            ws_reg.cell(ligne_reg, 35, data.classe_25)
                            ws_reg.cell(ligne_reg, 36, data.classe_26)
                            ws_reg.cell(ligne_reg, 37, data.classe_27)
                            ws_reg.cell(ligne_reg, 38, data.classe_28)
                            ws_reg.cell(ligne_reg, 39, data.classe_29)
                            ws_reg.cell(ligne_reg, 40, data.classe_30)
                            ws_reg.cell(ligne_reg, 41, data.classe_31)
                            ws_reg.cell(ligne_reg, 42, data.classe_32)
                            ws_reg.cell(ligne_reg, 43, data.classe_33)
                            ws_reg.cell(ligne_reg, 44, data.classe_34)
                            ws_reg.cell(ligne_reg, 45, data.classe_35)
                            ws_reg.cell(ligne_reg, 46, data.classe_36)
                            ws_reg.cell(ligne_reg, 47, data.classe_37)
                            ws_reg.cell(ligne_reg, 48, data.classe_38)
                            ws_reg.cell(ligne_reg, 49, data.classe_39)
                            ws_reg.cell(ligne_reg, 50, data.classe_40)
                            ws_reg.cell(ligne_reg, 51, data.classe_41)
                            ws_reg.cell(ligne_reg, 52, data.classe_42)
                            ws_reg.cell(ligne_reg, 53, data.classe_43)
                            ws_reg.cell(ligne_reg, 54, data.classe_44)
                            ws_reg.cell(ligne_reg, 55, data.classe_45)
                            ws_reg.cell(ligne_reg, 56, data.somme_classes)
                            ligne_reg += 1
                            numero_reg += 1
                            somme_reg += data.nombre_marques
                            total_reg += data.somme_classes
                            # Nationaux
                            if data.pays_marques.specification_pays == 'Nationaux':
                                ws_reg_nat.cell(ligne_nat, 1, numero_nat)
                                ws_reg_nat.cell(ligne_nat, 2, data.pays_marques.code_pays)
                                ws_reg_nat.cell(ligne_nat, 3, data.pays_marques.nom_pays)
                                ws_reg_nat.cell(ligne_nat, 4, data.nombre_marques)
                                ws_reg_nat.cell(ligne_nat, 5, data.somme_classes)

                                ws_reg_nat.cell(ligne_nat, 8, numero_nat)
                                ws_reg_nat.cell(ligne_nat, 9, data.pays_marques.code_pays)
                                ws_reg_nat.cell(ligne_nat, 10, data.pays_marques.nom_pays)
                                ws_reg_nat.cell(ligne_nat, 11, data.classe_1)
                                ws_reg_nat.cell(ligne_nat, 12, data.classe_2)
                                ws_reg_nat.cell(ligne_nat, 13, data.classe_3)
                                ws_reg_nat.cell(ligne_nat, 14, data.classe_4)
                                ws_reg_nat.cell(ligne_nat, 15, data.classe_5)
                                ws_reg_nat.cell(ligne_nat, 16, data.classe_6)
                                ws_reg_nat.cell(ligne_nat, 17, data.classe_7)
                                ws_reg_nat.cell(ligne_nat, 18, data.classe_8)
                                ws_reg_nat.cell(ligne_nat, 19, data.classe_9)
                                ws_reg_nat.cell(ligne_nat, 20, data.classe_10)
                                ws_reg_nat.cell(ligne_nat, 21, data.classe_11)
                                ws_reg_nat.cell(ligne_nat, 22, data.classe_12)
                                ws_reg_nat.cell(ligne_nat, 23, data.classe_13)
                                ws_reg_nat.cell(ligne_nat, 24, data.classe_14)
                                ws_reg_nat.cell(ligne_nat, 25, data.classe_15)
                                ws_reg_nat.cell(ligne_nat, 26, data.classe_16)
                                ws_reg_nat.cell(ligne_nat, 27, data.classe_17)
                                ws_reg_nat.cell(ligne_nat, 28, data.classe_18)
                                ws_reg_nat.cell(ligne_nat, 29, data.classe_19)
                                ws_reg_nat.cell(ligne_nat, 30, data.classe_20)
                                ws_reg_nat.cell(ligne_nat, 31, data.classe_21)
                                ws_reg_nat.cell(ligne_nat, 32, data.classe_22)
                                ws_reg_nat.cell(ligne_nat, 33, data.classe_23)
                                ws_reg_nat.cell(ligne_nat, 34, data.classe_24)
                                ws_reg_nat.cell(ligne_nat, 35, data.classe_25)
                                ws_reg_nat.cell(ligne_nat, 36, data.classe_26)
                                ws_reg_nat.cell(ligne_nat, 37, data.classe_27)
                                ws_reg_nat.cell(ligne_nat, 38, data.classe_28)
                                ws_reg_nat.cell(ligne_nat, 39, data.classe_29)
                                ws_reg_nat.cell(ligne_nat, 40, data.classe_30)
                                ws_reg_nat.cell(ligne_nat, 41, data.classe_31)
                                ws_reg_nat.cell(ligne_nat, 42, data.classe_32)
                                ws_reg_nat.cell(ligne_nat, 43, data.classe_33)
                                ws_reg_nat.cell(ligne_nat, 44, data.classe_34)
                                ws_reg_nat.cell(ligne_nat, 45, data.classe_35)
                                ws_reg_nat.cell(ligne_nat, 46, data.classe_36)
                                ws_reg_nat.cell(ligne_nat, 47, data.classe_37)
                                ws_reg_nat.cell(ligne_nat, 48, data.classe_38)
                                ws_reg_nat.cell(ligne_nat, 49, data.classe_39)
                                ws_reg_nat.cell(ligne_nat, 50, data.classe_40)
                                ws_reg_nat.cell(ligne_nat, 51, data.classe_41)
                                ws_reg_nat.cell(ligne_nat, 52, data.classe_42)
                                ws_reg_nat.cell(ligne_nat, 53, data.classe_43)
                                ws_reg_nat.cell(ligne_nat, 54, data.classe_44)
                                ws_reg_nat.cell(ligne_nat, 55, data.classe_45)
                                ws_reg_nat.cell(ligne_nat, 56, data.somme_classes)
                                ligne_nat += 1
                                numero_nat += 1
                                somme_nat += data.nombre_marques
                                total_nat += data.somme_classes
                            else:
                                # Etrangers
                                ws_reg_etran.cell(ligne_etran, 1, numero_etran)
                                ws_reg_etran.cell(ligne_etran, 2, data.pays_marques.code_pays)
                                ws_reg_etran.cell(ligne_etran, 3, data.pays_marques.nom_pays)
                                ws_reg_etran.cell(ligne_etran, 4, data.nombre_marques)
                                ws_reg_etran.cell(ligne_etran, 5, data.somme_classes)

                                ws_reg_etran.cell(ligne_etran, 8, numero_etran)
                                ws_reg_etran.cell(ligne_etran, 9, data.pays_marques.code_pays)
                                ws_reg_etran.cell(ligne_etran, 10, data.pays_marques.nom_pays)
                                ws_reg_etran.cell(ligne_etran, 11, data.classe_1)
                                ws_reg_etran.cell(ligne_etran, 12, data.classe_2)
                                ws_reg_etran.cell(ligne_etran, 13, data.classe_3)
                                ws_reg_etran.cell(ligne_etran, 14, data.classe_4)
                                ws_reg_etran.cell(ligne_etran, 15, data.classe_5)
                                ws_reg_etran.cell(ligne_etran, 16, data.classe_6)
                                ws_reg_etran.cell(ligne_etran, 17, data.classe_7)
                                ws_reg_etran.cell(ligne_etran, 18, data.classe_8)
                                ws_reg_etran.cell(ligne_etran, 19, data.classe_9)
                                ws_reg_etran.cell(ligne_etran, 20, data.classe_10)
                                ws_reg_etran.cell(ligne_etran, 21, data.classe_11)
                                ws_reg_etran.cell(ligne_etran, 22, data.classe_12)
                                ws_reg_etran.cell(ligne_etran, 23, data.classe_13)
                                ws_reg_etran.cell(ligne_etran, 24, data.classe_14)
                                ws_reg_etran.cell(ligne_etran, 25, data.classe_15)
                                ws_reg_etran.cell(ligne_etran, 26, data.classe_16)
                                ws_reg_etran.cell(ligne_etran, 27, data.classe_17)
                                ws_reg_etran.cell(ligne_etran, 28, data.classe_18)
                                ws_reg_etran.cell(ligne_etran, 29, data.classe_19)
                                ws_reg_etran.cell(ligne_etran, 30, data.classe_20)
                                ws_reg_etran.cell(ligne_etran, 31, data.classe_21)
                                ws_reg_etran.cell(ligne_etran, 32, data.classe_22)
                                ws_reg_etran.cell(ligne_etran, 33, data.classe_23)
                                ws_reg_etran.cell(ligne_etran, 34, data.classe_24)
                                ws_reg_etran.cell(ligne_etran, 35, data.classe_25)
                                ws_reg_etran.cell(ligne_etran, 36, data.classe_26)
                                ws_reg_etran.cell(ligne_etran, 37, data.classe_27)
                                ws_reg_etran.cell(ligne_etran, 38, data.classe_28)
                                ws_reg_etran.cell(ligne_etran, 39, data.classe_29)
                                ws_reg_etran.cell(ligne_etran, 40, data.classe_30)
                                ws_reg_etran.cell(ligne_etran, 41, data.classe_31)
                                ws_reg_etran.cell(ligne_etran, 42, data.classe_32)
                                ws_reg_etran.cell(ligne_etran, 43, data.classe_33)
                                ws_reg_etran.cell(ligne_etran, 44, data.classe_34)
                                ws_reg_etran.cell(ligne_etran, 45, data.classe_35)
                                ws_reg_etran.cell(ligne_etran, 46, data.classe_36)
                                ws_reg_etran.cell(ligne_etran, 47, data.classe_37)
                                ws_reg_etran.cell(ligne_etran, 48, data.classe_38)
                                ws_reg_etran.cell(ligne_etran, 49, data.classe_39)
                                ws_reg_etran.cell(ligne_etran, 50, data.classe_40)
                                ws_reg_etran.cell(ligne_etran, 51, data.classe_41)
                                ws_reg_etran.cell(ligne_etran, 52, data.classe_42)
                                ws_reg_etran.cell(ligne_etran, 53, data.classe_43)
                                ws_reg_etran.cell(ligne_etran, 54, data.classe_44)
                                ws_reg_etran.cell(ligne_etran, 55, data.classe_45)
                                ws_reg_etran.cell(ligne_etran, 56, data.somme_classes)
                                ligne_etran += 1
                                numero_etran += 1
                                somme_etran += data.nombre_marques
                                total_etran += data.somme_classes
                        # Madrid
                        if "Madrid" in data.type_marques:
                            ws_madrid.cell(ligne_madrid, 1, numero_madrid)
                            ws_madrid.cell(ligne_madrid, 2, data.pays_marques.code_pays)
                            ws_madrid.cell(ligne_madrid, 3, data.pays_marques.nom_pays)
                            ws_madrid.cell(ligne_madrid, 4, data.nombre_marques)
                            ws_madrid.cell(ligne_madrid, 5, data.somme_classes)

                            ws_madrid.cell(ligne_madrid, 8, numero_madrid)
                            ws_madrid.cell(ligne_madrid, 9, data.pays_marques.code_pays)
                            ws_madrid.cell(ligne_madrid, 10, data.pays_marques.nom_pays)
                            ws_madrid.cell(ligne_madrid, 11, data.classe_1)
                            ws_madrid.cell(ligne_madrid, 12, data.classe_2)
                            ws_madrid.cell(ligne_madrid, 13, data.classe_3)
                            ws_madrid.cell(ligne_madrid, 14, data.classe_4)
                            ws_madrid.cell(ligne_madrid, 15, data.classe_5)
                            ws_madrid.cell(ligne_madrid, 16, data.classe_6)
                            ws_madrid.cell(ligne_madrid, 17, data.classe_7)
                            ws_madrid.cell(ligne_madrid, 18, data.classe_8)
                            ws_madrid.cell(ligne_madrid, 19, data.classe_9)
                            ws_madrid.cell(ligne_madrid, 20, data.classe_10)
                            ws_madrid.cell(ligne_madrid, 21, data.classe_11)
                            ws_madrid.cell(ligne_madrid, 22, data.classe_12)
                            ws_madrid.cell(ligne_madrid, 23, data.classe_13)
                            ws_madrid.cell(ligne_madrid, 24, data.classe_14)
                            ws_madrid.cell(ligne_madrid, 25, data.classe_15)
                            ws_madrid.cell(ligne_madrid, 26, data.classe_16)
                            ws_madrid.cell(ligne_madrid, 27, data.classe_17)
                            ws_madrid.cell(ligne_madrid, 28, data.classe_18)
                            ws_madrid.cell(ligne_madrid, 29, data.classe_19)
                            ws_madrid.cell(ligne_madrid, 30, data.classe_20)
                            ws_madrid.cell(ligne_madrid, 31, data.classe_21)
                            ws_madrid.cell(ligne_madrid, 32, data.classe_22)
                            ws_madrid.cell(ligne_madrid, 33, data.classe_23)
                            ws_madrid.cell(ligne_madrid, 34, data.classe_24)
                            ws_madrid.cell(ligne_madrid, 35, data.classe_25)
                            ws_madrid.cell(ligne_madrid, 36, data.classe_26)
                            ws_madrid.cell(ligne_madrid, 37, data.classe_27)
                            ws_madrid.cell(ligne_madrid, 38, data.classe_28)
                            ws_madrid.cell(ligne_madrid, 39, data.classe_29)
                            ws_madrid.cell(ligne_madrid, 40, data.classe_30)
                            ws_madrid.cell(ligne_madrid, 41, data.classe_31)
                            ws_madrid.cell(ligne_madrid, 42, data.classe_32)
                            ws_madrid.cell(ligne_madrid, 43, data.classe_33)
                            ws_madrid.cell(ligne_madrid, 44, data.classe_34)
                            ws_madrid.cell(ligne_madrid, 45, data.classe_35)
                            ws_madrid.cell(ligne_madrid, 46, data.classe_36)
                            ws_madrid.cell(ligne_madrid, 47, data.classe_37)
                            ws_madrid.cell(ligne_madrid, 48, data.classe_38)
                            ws_madrid.cell(ligne_madrid, 49, data.classe_39)
                            ws_madrid.cell(ligne_madrid, 50, data.classe_40)
                            ws_madrid.cell(ligne_madrid, 51, data.classe_41)
                            ws_madrid.cell(ligne_madrid, 52, data.classe_42)
                            ws_madrid.cell(ligne_madrid, 53, data.classe_43)
                            ws_madrid.cell(ligne_madrid, 54, data.classe_44)
                            ws_madrid.cell(ligne_madrid, 55, data.classe_45)
                            ws_madrid.cell(ligne_madrid, 56, data.somme_classes)
                            ligne_madrid += 1
                            numero_madrid += 1
                            somme_madrid += data.nombre_marques
                            total_madrid += data.somme_classes
                    # Voie Globale
                    for p in Pays.objects.all().order_by('nom_pays'):
                        # Tous les pays
                        if type_pi == 'Demande':
                            data = Marques.objects.filter(Q(pays_marques=p.id) & Q(type_marques__contains=type_pi)
                                                          & Q(annee_marques=annee))
                        else:
                            data = Marques.objects.filter(Q(pays_marques=p.id) & Q(type_marques__contains='Enreg')
                                                          & Q(annee_marques=annee))
                        if data.exists():
                            ws.cell(ligne, 1, numero)
                            ws.cell(ligne, 2, data.first().pays_marques.code_pays)
                            ws.cell(ligne, 3, data.first().pays_marques.nom_pays)
                            ws.cell(ligne, 4,
                                    data.aggregate(Sum('nombre_marques'))['nombre_marques__sum'])
                            ws.cell(ligne, 5, sum([d.somme_classes for d in data]))

                            ws.cell(ligne, 8, numero)
                            ws.cell(ligne, 9, data.first().pays_marques.code_pays)
                            ws.cell(ligne, 10, data.first().pays_marques.nom_pays)
                            ws.cell(ligne, 11, data.aggregate(Sum('classe_1'))['classe_1__sum'])
                            ws.cell(ligne, 12, data.aggregate(Sum('classe_2'))['classe_2__sum'])
                            ws.cell(ligne, 13, data.aggregate(Sum('classe_3'))['classe_3__sum'])
                            ws.cell(ligne, 14, data.aggregate(Sum('classe_4'))['classe_4__sum'])
                            ws.cell(ligne, 15, data.aggregate(Sum('classe_5'))['classe_5__sum'])
                            ws.cell(ligne, 16, data.aggregate(Sum('classe_6'))['classe_6__sum'])
                            ws.cell(ligne, 17, data.aggregate(Sum('classe_7'))['classe_7__sum'])
                            ws.cell(ligne, 18, data.aggregate(Sum('classe_8'))['classe_8__sum'])
                            ws.cell(ligne, 19, data.aggregate(Sum('classe_9'))['classe_9__sum'])
                            ws.cell(ligne, 20, data.aggregate(Sum('classe_10'))['classe_10__sum'])
                            ws.cell(ligne, 21, data.aggregate(Sum('classe_11'))['classe_11__sum'])
                            ws.cell(ligne, 22, data.aggregate(Sum('classe_12'))['classe_12__sum'])
                            ws.cell(ligne, 23, data.aggregate(Sum('classe_13'))['classe_13__sum'])
                            ws.cell(ligne, 24, data.aggregate(Sum('classe_14'))['classe_14__sum'])
                            ws.cell(ligne, 25, data.aggregate(Sum('classe_15'))['classe_15__sum'])
                            ws.cell(ligne, 26, data.aggregate(Sum('classe_16'))['classe_16__sum'])
                            ws.cell(ligne, 27, data.aggregate(Sum('classe_17'))['classe_17__sum'])
                            ws.cell(ligne, 28, data.aggregate(Sum('classe_18'))['classe_18__sum'])
                            ws.cell(ligne, 29, data.aggregate(Sum('classe_19'))['classe_19__sum'])
                            ws.cell(ligne, 30, data.aggregate(Sum('classe_20'))['classe_20__sum'])
                            ws.cell(ligne, 31, data.aggregate(Sum('classe_21'))['classe_21__sum'])
                            ws.cell(ligne, 32, data.aggregate(Sum('classe_22'))['classe_22__sum'])
                            ws.cell(ligne, 33, data.aggregate(Sum('classe_23'))['classe_23__sum'])
                            ws.cell(ligne, 34, data.aggregate(Sum('classe_24'))['classe_24__sum'])
                            ws.cell(ligne, 35, data.aggregate(Sum('classe_25'))['classe_25__sum'])
                            ws.cell(ligne, 36, data.aggregate(Sum('classe_26'))['classe_26__sum'])
                            ws.cell(ligne, 37, data.aggregate(Sum('classe_27'))['classe_27__sum'])
                            ws.cell(ligne, 38, data.aggregate(Sum('classe_28'))['classe_28__sum'])
                            ws.cell(ligne, 39, data.aggregate(Sum('classe_29'))['classe_29__sum'])
                            ws.cell(ligne, 40, data.aggregate(Sum('classe_30'))['classe_30__sum'])
                            ws.cell(ligne, 41, data.aggregate(Sum('classe_31'))['classe_31__sum'])
                            ws.cell(ligne, 42, data.aggregate(Sum('classe_32'))['classe_32__sum'])
                            ws.cell(ligne, 43, data.aggregate(Sum('classe_33'))['classe_33__sum'])
                            ws.cell(ligne, 44, data.aggregate(Sum('classe_34'))['classe_34__sum'])
                            ws.cell(ligne, 45, data.aggregate(Sum('classe_35'))['classe_35__sum'])
                            ws.cell(ligne, 46, data.aggregate(Sum('classe_36'))['classe_36__sum'])
                            ws.cell(ligne, 47, data.aggregate(Sum('classe_37'))['classe_37__sum'])
                            ws.cell(ligne, 48, data.aggregate(Sum('classe_38'))['classe_38__sum'])
                            ws.cell(ligne, 49, data.aggregate(Sum('classe_39'))['classe_39__sum'])
                            ws.cell(ligne, 50, data.aggregate(Sum('classe_40'))['classe_40__sum'])
                            ws.cell(ligne, 51, data.aggregate(Sum('classe_41'))['classe_41__sum'])
                            ws.cell(ligne, 52, data.aggregate(Sum('classe_42'))['classe_42__sum'])
                            ws.cell(ligne, 53, data.aggregate(Sum('classe_43'))['classe_43__sum'])
                            ws.cell(ligne, 54, data.aggregate(Sum('classe_44'))['classe_44__sum'])
                            ws.cell(ligne, 55, data.aggregate(Sum('classe_45'))['classe_45__sum'])
                            ws.cell(ligne, 56, sum([d.somme_classes for d in data]))
                            ligne += 1
                            numero += 1
                            somme += data.aggregate(Sum('nombre_marques'))['nombre_marques__sum']
                            total += sum([d.somme_classes for d in data])
                        # Etrangers
                        if p.specification_pays == 'Etrangers':
                            if type_pi == 'Demande':
                                data = Marques.objects.filter(Q(pays_marques=p.id) & Q(type_marques__contains=type_pi)
                                                              & Q(annee_marques=annee))
                            else:
                                data = Marques.objects.filter(Q(pays_marques=p.id) & Q(type_marques__contains='Enreg')
                                                              & Q(annee_marques=annee))
                            if data.exists():
                                ws_etran_global.cell(ligne_etr_glob, 1, numero_etr_glob)
                                ws_etran_global.cell(ligne_etr_glob, 2, data.first().pays_marques.code_pays)
                                ws_etran_global.cell(ligne_etr_glob, 3, data.first().pays_marques.nom_pays)
                                ws_etran_global.cell(ligne_etr_glob, 4,
                                                     data.aggregate(Sum('nombre_marques'))['nombre_marques__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 5, sum([d.somme_classes for d in data]))

                                ws_etran_global.cell(ligne_etr_glob, 8, numero_etr_glob)
                                ws_etran_global.cell(ligne_etr_glob, 9, data.first().pays_marques.code_pays)
                                ws_etran_global.cell(ligne_etr_glob, 10, data.first().pays_marques.nom_pays)
                                ws_etran_global.cell(ligne_etr_glob, 11,
                                                     data.aggregate(Sum('classe_1'))['classe_1__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 12,
                                                     data.aggregate(Sum('classe_2'))['classe_2__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 13,
                                                     data.aggregate(Sum('classe_3'))['classe_3__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 14,
                                                     data.aggregate(Sum('classe_4'))['classe_4__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 15,
                                                     data.aggregate(Sum('classe_5'))['classe_5__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 16,
                                                     data.aggregate(Sum('classe_6'))['classe_6__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 17,
                                                     data.aggregate(Sum('classe_7'))['classe_7__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 18,
                                                     data.aggregate(Sum('classe_8'))['classe_8__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 19,
                                                     data.aggregate(Sum('classe_9'))['classe_9__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 20,
                                                     data.aggregate(Sum('classe_10'))['classe_10__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 21,
                                                     data.aggregate(Sum('classe_11'))['classe_11__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 22,
                                                     data.aggregate(Sum('classe_12'))['classe_12__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 23,
                                                     data.aggregate(Sum('classe_13'))['classe_13__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 24,
                                                     data.aggregate(Sum('classe_14'))['classe_14__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 25,
                                                     data.aggregate(Sum('classe_15'))['classe_15__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 26,
                                                     data.aggregate(Sum('classe_16'))['classe_16__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 27,
                                                     data.aggregate(Sum('classe_17'))['classe_17__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 28,
                                                     data.aggregate(Sum('classe_18'))['classe_18__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 29,
                                                     data.aggregate(Sum('classe_19'))['classe_19__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 30,
                                                     data.aggregate(Sum('classe_20'))['classe_20__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 31,
                                                     data.aggregate(Sum('classe_21'))['classe_21__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 32,
                                                     data.aggregate(Sum('classe_22'))['classe_22__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 33,
                                                     data.aggregate(Sum('classe_23'))['classe_23__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 34,
                                                     data.aggregate(Sum('classe_24'))['classe_24__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 35,
                                                     data.aggregate(Sum('classe_25'))['classe_25__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 36,
                                                     data.aggregate(Sum('classe_26'))['classe_26__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 37,
                                                     data.aggregate(Sum('classe_27'))['classe_27__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 38,
                                                     data.aggregate(Sum('classe_28'))['classe_28__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 39,
                                                     data.aggregate(Sum('classe_29'))['classe_29__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 40,
                                                     data.aggregate(Sum('classe_30'))['classe_30__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 41,
                                                     data.aggregate(Sum('classe_31'))['classe_31__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 42,
                                                     data.aggregate(Sum('classe_32'))['classe_32__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 43,
                                                     data.aggregate(Sum('classe_33'))['classe_33__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 44,
                                                     data.aggregate(Sum('classe_34'))['classe_34__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 45,
                                                     data.aggregate(Sum('classe_35'))['classe_35__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 46,
                                                     data.aggregate(Sum('classe_36'))['classe_36__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 47,
                                                     data.aggregate(Sum('classe_37'))['classe_37__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 48,
                                                     data.aggregate(Sum('classe_38'))['classe_38__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 49,
                                                     data.aggregate(Sum('classe_39'))['classe_39__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 50,
                                                     data.aggregate(Sum('classe_40'))['classe_40__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 51,
                                                     data.aggregate(Sum('classe_41'))['classe_41__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 52,
                                                     data.aggregate(Sum('classe_42'))['classe_42__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 53,
                                                     data.aggregate(Sum('classe_43'))['classe_43__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 54,
                                                     data.aggregate(Sum('classe_44'))['classe_44__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 55,
                                                     data.aggregate(Sum('classe_45'))['classe_45__sum'])
                                ws_etran_global.cell(ligne_etr_glob, 56, sum([d.somme_classes for d in data]))
                                ligne_etr_glob += 1
                                numero_etr_glob += 1
                                somme_etr_glob += data.aggregate(Sum('nombre_marques'))['nombre_marques__sum']
                                total_etr_glob += sum([d.somme_classes for d in data])
                    # Totaux des fichiers
                    # Voie régionale
                    ws_reg.cell(ligne_reg, 4, somme_reg)
                    ws_reg.cell(ligne_reg, 5, total_reg)
                    # Voie régionale nationale
                    ws_reg_nat.cell(ligne_nat, 4, somme_nat)
                    ws_reg_nat.cell(ligne_nat, 5, total_nat)
                    # Voie régionale etrangers
                    ws_reg_etran.cell(ligne_etran, 4, somme_etran)
                    ws_reg_etran.cell(ligne_etran, 5, total_etran)
                    # Voie Madrid
                    ws_madrid.cell(ligne_madrid, 4, somme_madrid)
                    ws_madrid.cell(ligne_madrid, 5, total_madrid)
                    # Global Etrangers
                    ws_etran_global.cell(ligne_etr_glob, 4, somme_etr_glob)
                    ws_etran_global.cell(ligne_etr_glob, 5, total_etr_glob)
                    # Global
                    ws.cell(ligne, 4, somme)
                    ws.cell(ligne, 5, total)
                    # Sauvergarde du fichier
                    wb.save(settings.MEDIA_ROOT + "/" + pi + " - " + type_pi + " " + str(annee) + ".xlsx")
                    file = File(
                        open(settings.MEDIA_ROOT + "/" + pi + " - " + type_pi + " " + str(annee) + ".xlsx", mode='rb'),
                        name=pi + ' - ' + type_pi + ' ' + str(annee) + ".xlsx")
                    messages.success(request, "Les données ont été exportées sous forme de fichier excel.",
                                     extra_tags='success')
                    # On enregistre l'historique
                    h = Historique.objects.create(type_fichier="OAPI", nature_fichier=type_pi, type_pi=pi,
                                                  upload_file=file, annee=annee)
                    h.save()
                    return redirect('core:export_file_oapi')
                else:
                    messages.warning(request, "Il n'existe aucun {0} de type {1} pour l'année {2} dans la base de"
                                              " données.".format(pi, type_pi, annee), extra_tags='warning')
                    return redirect('core:export_file_oapi')
        else:
            messages.error(request, "Veuillez remplir tous les champs du formulaire s'il vous plaît.",
                           extra_tags='danger')
            return redirect('core:export_file_oapi')

    context = {
        'menu': menu,
        'i_menu': i_menu,
        'history': history
    }
    return render(request, 'core/export_file_oapi.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def export_file_ompi(request):
    # Active menu
    menu = 'config'
    i_menu = 'ompi'
    extensions = ['xls', 'xlsx']
    # path_file = "C:\\Users\\stefn\\OneDrive\\Bureau\\Fichiers OAPI\\"

    # On récupère l'historique
    history = Historique.objects.all().order_by('-date_ajout')

    if request.method == 'POST':
        form = ExportEFileForm(request.POST, request.FILES)
        # Formulaire valide
        if form.is_valid():
            pi = form.cleaned_data['pi']
            type_pi = form.cleaned_data['type_pi']
            annee = form.cleaned_data['annee']
            file = form.cleaned_data['file']
            # S'il s'agit d'un fichier excel
            if str(file).split('.')[1] in extensions:
                wb = load_workbook(filename=file)
                ws = wb.active
                # Brevets
                if pi == 'Brevets':
                    # On recupere les données
                    data = get_ompi_data(pi, type_pi, annee)
                    if len(data) > 0:
                        # Remplissage de cellules
                        for row in ws:
                            if row[1].value is not None:
                                if row[1].value in data:
                                    row[3].value = data[row[1].value][0]
                                    row[4].value = data[row[1].value][1]
                                else:
                                    if len(row[1].value) < 3:
                                        row[3].value = 0
                                        row[4].value = 0
                        wb.save(file)
                        messages.success(request, "Les données ont été exportées dans le fichier {}".format(str(file)),
                                         extra_tags='success')
                        # On enregistre l'historique
                        h = Historique.objects.create(type_fichier="OMPI", nature_fichier=type_pi, type_pi=pi,
                                                       upload_file=file, annee=annee)
                        h.save()
                        return redirect('core:export_file_ompi')
                    else:
                        messages.warning(request, "Aucune données pour l'année {}. Veuillez importer des données "
                                                  "s'il vous plaît.".format(annee), extra_tags='warning')
                        return redirect('core:load_brevets')
                # DMI
                if pi == 'DMI Origin':
                    # On recupere les données
                    data = get_ompi_data(pi, type_pi, annee)
                    if len(data) > 0:
                        # Remplissage de cellules
                        for row in ws:
                            if row[1].value is not None:
                                if row[1].value in data:
                                    row[3].value = data[row[1].value][0]
                                    row[4].value = data[row[1].value][1]
                                else:
                                    if len(row[1].value) < 3:
                                        row[3].value = 0
                                        row[4].value = 0
                        wb.save(file)
                        messages.success(request, "Les données ont été exportées dans le fichier {}".format(str(file)),
                                         extra_tags='success')
                        # On enregistre l'historique
                        h = Historique.objects.create(type_fichier="OMPI", nature_fichier=type_pi,
                                                      type_pi=pi + " Origin", upload_file=file,
                                                      annee=annee)
                        h.save()
                        return redirect('core:export_file_ompi')
                    else:
                        messages.warning(request, "Aucune données pour l'année {}. Veuillez importer des données "
                                                  "s'il vous plaît.".format(annee), extra_tags='warning')
                        return redirect('core:load_dmi')
                # DMI Classes
                if pi == 'DMI Classes':
                    # On recupere les données
                    data = get_ompi_data(pi, type_pi, annee)
                    if len(data) > 0:
                        # Remplissage de cellules
                        for row in ws:
                            if row[1].value is not None:
                                if row[1].value in data:
                                    row[3].value = data[row[1].value][0]
                                    row[4].value = data[row[1].value][1]
                                    row[5].value = data[row[1].value][2]
                                    row[6].value = data[row[1].value][3]
                                    row[7].value = data[row[1].value][4]
                                    row[8].value = data[row[1].value][5]
                                    row[9].value = data[row[1].value][6]
                                    row[10].value = data[row[1].value][7]
                                    row[11].value = data[row[1].value][8]
                                    row[12].value = data[row[1].value][9]
                                    row[13].value = data[row[1].value][10]
                                    row[14].value = data[row[1].value][11]
                                    row[15].value = data[row[1].value][12]
                                    row[16].value = data[row[1].value][13]
                                    row[17].value = data[row[1].value][14]
                                    row[18].value = data[row[1].value][15]
                                    row[19].value = data[row[1].value][16]
                                    row[20].value = data[row[1].value][17]
                                    row[21].value = data[row[1].value][18]
                                    row[22].value = data[row[1].value][19]
                                    row[23].value = data[row[1].value][20]
                                    row[24].value = data[row[1].value][21]
                                    row[25].value = data[row[1].value][22]
                                    row[26].value = data[row[1].value][23]
                                    row[27].value = data[row[1].value][24]
                                    row[28].value = data[row[1].value][25]
                                    row[29].value = data[row[1].value][26]
                                    row[30].value = data[row[1].value][27]
                                    row[31].value = data[row[1].value][28]
                                    row[32].value = data[row[1].value][29]
                                    row[33].value = data[row[1].value][30]
                                    row[34].value = data[row[1].value][31]
                                    row[35].value = data[row[1].value][32]
                                    row[36].value = data[row[1].value][33]
                                else:
                                    if len(row[1].value) < 3:
                                        row[3].value = 0
                                        row[4].value = 0
                                        row[5].value = 0
                                        row[6].value = 0
                                        row[7].value = 0
                                        row[8].value = 0
                                        row[9].value = 0
                                        row[10].value = 0
                                        row[11].value = 0
                                        row[12].value = 0
                                        row[13].value = 0
                                        row[14].value = 0
                                        row[15].value = 0
                                        row[16].value = 0
                                        row[17].value = 0
                                        row[18].value = 0
                                        row[19].value = 0
                                        row[20].value = 0
                                        row[21].value = 0
                                        row[22].value = 0
                                        row[23].value = 0
                                        row[24].value = 0
                                        row[25].value = 0
                                        row[26].value = 0
                                        row[27].value = 0
                                        row[28].value = 0
                                        row[29].value = 0
                                        row[30].value = 0
                                        row[31].value = 0
                                        row[32].value = 0
                                        row[33].value = 0
                                        row[34].value = 0
                                        row[35].value = 0
                                        row[36].value = 0
                        wb.save(file)
                        messages.success(request, "Les données ont été exportées dans le fichier {}".format(str(file)),
                                         extra_tags='success')
                        # On enregistre l'historique
                        h = Historique.objects.create(type_fichier="OMPI", nature_fichier=type_pi + " Classes",
                                                      type_pi=pi, upload_file=file,
                                                      annee=annee)
                        h.save()
                        return redirect('core:export_file_ompi')
                    else:
                        messages.warning(request, "Aucune données pour l'année {}. Veuillez importer des données "
                                                  "s'il vous plaît.".format(annee), extra_tags='warning')
                        return redirect('core:load_dmi')
                # Marques
                if pi == 'Marque Origin':
                    # On recupere les données
                    data = get_ompi_data(pi, type_pi, annee)
                    if len(data) > 0:
                        # Remplissage de cellules
                        for row in ws:
                            if row[1].value is not None:
                                if row[1].value in data:
                                    row[3].value = data[row[1].value][0]
                                    row[4].value = data[row[1].value][1]
                                else:
                                    if len(row[1].value) < 3:
                                        row[3].value = 0
                                        row[4].value = 0
                        wb.save(file)
                        messages.success(request,
                                         "Les données ont été exportées dans le fichier {}".format(str(file)),
                                         extra_tags='success')
                        # On enregistre l'historique
                        h = Historique.objects.create(type_fichier="OMPI", nature_fichier=type_pi,
                                                      type_pi=pi + " Origin", upload_file=file,
                                                      annee=annee)
                        h.save()
                        return redirect('core:export_file_ompi')
                    else:
                        messages.warning(request,
                                         "Aucune données pour l'année {}. Veuillez importer des données "
                                         "s'il vous plaît.".format(annee), extra_tags='warning')
                        return redirect('core:load_marques')
                # Marques Classes
                if pi == 'Marque Classes':
                    # On recupere les données
                    data = get_ompi_data(pi, type_pi, annee)
                    if len(data) > 0:
                        # Remplissage de cellules
                        for row in ws:
                            if row[1].value is not None:
                                if row[1].value in data:
                                    row[3].value = data[row[1].value][0]
                                    row[4].value = data[row[1].value][1]
                                    row[5].value = data[row[1].value][2]
                                    row[6].value = data[row[1].value][3]
                                    row[7].value = data[row[1].value][4]
                                    row[8].value = data[row[1].value][5]
                                    row[9].value = data[row[1].value][6]
                                    row[10].value = data[row[1].value][7]
                                    row[11].value = data[row[1].value][8]
                                    row[12].value = data[row[1].value][9]
                                    row[13].value = data[row[1].value][10]
                                    row[14].value = data[row[1].value][11]
                                    row[15].value = data[row[1].value][12]
                                    row[16].value = data[row[1].value][13]
                                    row[17].value = data[row[1].value][14]
                                    row[18].value = data[row[1].value][15]
                                    row[19].value = data[row[1].value][16]
                                    row[20].value = data[row[1].value][17]
                                    row[21].value = data[row[1].value][18]
                                    row[22].value = data[row[1].value][19]
                                    row[23].value = data[row[1].value][20]
                                    row[24].value = data[row[1].value][21]
                                    row[25].value = data[row[1].value][22]
                                    row[26].value = data[row[1].value][23]
                                    row[27].value = data[row[1].value][24]
                                    row[28].value = data[row[1].value][25]
                                    row[29].value = data[row[1].value][26]
                                    row[30].value = data[row[1].value][27]
                                    row[31].value = data[row[1].value][28]
                                    row[32].value = data[row[1].value][29]
                                    row[33].value = data[row[1].value][30]
                                    row[34].value = data[row[1].value][31]
                                    row[35].value = data[row[1].value][32]
                                    row[36].value = data[row[1].value][33]
                                    row[37].value = data[row[1].value][34]
                                    row[38].value = data[row[1].value][35]
                                    row[39].value = data[row[1].value][36]
                                    row[40].value = data[row[1].value][37]
                                    row[41].value = data[row[1].value][38]
                                    row[42].value = data[row[1].value][39]
                                    row[43].value = data[row[1].value][40]
                                    row[44].value = data[row[1].value][41]
                                    row[45].value = data[row[1].value][42]
                                    row[46].value = data[row[1].value][43]
                                    row[47].value = data[row[1].value][44]
                                    row[48].value = data[row[1].value][45]
                                    row[49].value = data[row[1].value][46]
                                else:
                                    if len(row[1].value) < 3:
                                        row[3].value = 0
                                        row[4].value = 0
                                        row[5].value = 0
                                        row[6].value = 0
                                        row[7].value = 0
                                        row[8].value = 0
                                        row[9].value = 0
                                        row[10].value = 0
                                        row[11].value = 0
                                        row[12].value = 0
                                        row[13].value = 0
                                        row[14].value = 0
                                        row[15].value = 0
                                        row[16].value = 0
                                        row[17].value = 0
                                        row[18].value = 0
                                        row[19].value = 0
                                        row[20].value = 0
                                        row[21].value = 0
                                        row[22].value = 0
                                        row[23].value = 0
                                        row[24].value = 0
                                        row[25].value = 0
                                        row[26].value = 0
                                        row[27].value = 0
                                        row[28].value = 0
                                        row[29].value = 0
                                        row[30].value = 0
                                        row[31].value = 0
                                        row[32].value = 0
                                        row[33].value = 0
                                        row[34].value = 0
                                        row[35].value = 0
                                        row[36].value = 0
                                        row[37].value = 0
                                        row[38].value = 0
                                        row[39].value = 0
                                        row[40].value = 0
                                        row[41].value = 0
                                        row[42].value = 0
                                        row[43].value = 0
                                        row[44].value = 0
                                        row[45].value = 0
                                        row[46].value = 0
                                        row[47].value = 0
                                        row[48].value = 0
                                        row[49].value = 0
                        wb.save(file)
                        messages.success(request,
                                         "Les données ont été exportées dans le fichier {}".format(str(file)),
                                         extra_tags='success')
                        # On enregistre l'historique
                        h = Historique.objects.create(type_fichier="OMPI", nature_fichier=type_pi,
                                                      type_pi=pi + " Classes", upload_file=file,
                                                      annee=annee)
                        h.save()
                        return redirect('core:export_file_ompi')
                    else:
                        messages.warning(request,
                                         "Aucune données pour l'année {}. Veuillez importer des données "
                                         "s'il vous plaît.".format(annee), extra_tags='warning')
                        return redirect('core:load_marques')
            else:
                messages.warning(request, "Le fichier sélectionné doit être de format excel.", extra_tags='warning')
                return redirect('core:export_file_ompi')
        else:
            messages.error(request, "Veuillez remplir tous les champs du formulaire s'il vous plaît.",
                           extra_tags='danger')
            return redirect('core:export_file_ompi')

    context = {
        'menu': menu,
        'i_menu': i_menu,
        'history': history
    }
    return render(request, 'core/export_file_ompi.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def check_file_pays(request):
    # Active menu
    menu = 'pays'
    extensions = ['xls', 'xlsx']
    file_sheets_names = ['Nationaux', 'Etrangers']
    file_name = "Pays"
    error_names = 2
    total_rows = 0

    if request.method == 'POST':
        form = UploadPaysFileForm(request.POST or None, request.FILES or None)
        if form.is_valid():
            # On recupere les données du formulaire
            file = form.cleaned_data['file']
            # On vérifie le nom du fichier
            if file_name in str(file).split(".")[0]:
                # On vérifie l'extensions
                if str(file).split(".")[1] in extensions:
                    # Lecture du fichier
                    book = load_workbook(file)
                    # On vérifie si les feuilles du classeur sont conforme
                    for n in file_sheets_names:
                        if n not in book.sheetnames:
                            error_names -= 1
                    # Messages d'erreur si les feuilles ne sont pas conforment
                    if error_names == 0:
                        messages.warning(request, "Certains noms de feuilles du classeur sont incorrectes, veuillez "
                                                  "modifier les noms par ceux qui correspondent.", extra_tags='warning')
                        return redirect('core:load_pays')
                    else:
                        # Insertion de données pour chaque feuille
                        for n in file_sheets_names:
                            for row in book[n]:
                                if row[1].value is not None:
                                    if row[1].value != "Code":
                                        if not Pays.objects.filter(Q(code_pays=row[1].value)).exists():
                                            Pays.objects.create(
                                                code_pays=row[1].value,
                                                nom_pays=row[2].value,
                                                specification_pays=n
                                            )
                                            total_rows += 1
                        # Insertion réussie
                        if total_rows > 0:
                            messages.success(request, "{} pays ajouté dans la base de données.".format(total_rows),
                                             extra_tags='success')
                            return redirect('core:listing_pays')
                        else:
                            messages.info(request, "Aucun Pays ajouté dans la base de données.", extra_tags='info')
                            return redirect('core:listing_pays')
                else:
                    messages.warning(request, "Veuillez choisir comme fichier un classeur excel s'il vous plaît.",
                                     extra_tags='warning')
                    return redirect('core:load_pays')
            else:
                messages.warning(request, "Veuillez sélectionner un fichier de Pays s'il vous plaît.",
                                 extra_tags='warning')
                return redirect('core:load_pays')
        else:
            messages.error(request, "Le formulaire semble mal rempli, veuillez contacter l'administrateur système ou"
                                    " le statisticien si le problème persiste.", extra_tags='danger')
            return redirect('core:load_pays')

    context = {
        'menu': menu
    }
    return render(request, 'core/load_pays.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def check_file_brevets(request):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "brevets"
    extensions = ['xls', 'xlsx']
    file_sheets_names = ['Nationaux', 'Etrangers']
    file_name = "Brevets"
    error_names = 2
    total_rows = 0

    if request.method == 'POST':
        form = UploadPIFileForm(request.POST or None, request.FILES or None)
        if form.is_valid():
            # On recupere les données du formulaire
            file = form.cleaned_data['file']
            type_pi = form.cleaned_data['type_pi']
            annee = form.cleaned_data['annee']
            # Année valide
            if annee > 1964:
                # nom du fichier correct
                if file_name in str(file).split(".")[0]:
                    # Extensions correcte
                    if str(file).split(".")[1] in extensions:
                        # Lecture du fichier
                        book = load_workbook(file)
                        # On vérifie si les feuilles du classeur sont conforme
                        for n in file_sheets_names:
                            if n not in book.sheetnames:
                                error_names -= 1
                        # Messages d'erreur si les feuilles ne sont pas conforment
                        if error_names == 0:
                            messages.warning(request,
                                             "Certains noms de feuilles du classeur sont incorrectes, veuillez "
                                             "modifier les noms par ceux qui correspondent.", extra_tags='warning')
                            return redirect('core:load_brevets')
                        else:
                            # Insertion de données pour chaque feuille
                            for n in file_sheets_names:
                                for row in book[n]:
                                    if row[1].value is not None:
                                        # Si le pays n'existe pas on le crée
                                        if not Pays.objects.filter(Q(code_pays=row[1].value)).exists():
                                            Pays.objects.create(
                                                code_pays=row[1].value,
                                                nom_pays=row[2].value,
                                                specification_pays=n
                                            )

                                        # on recupère le pays
                                        p = Pays.objects.get(code_pays=row[1].value)

                                        # Si le brevet n'existe pas
                                        if not Brevets.objects.filter(Q(type_brevets=type_pi)
                                                                      & Q(annee_brevets=annee)
                                                                      & Q(pays_brevets=p)).exists():
                                            # on crée le brevet
                                            Brevets.objects.create(
                                                type_brevets=type_pi,
                                                annee_brevets=annee,
                                                nombre_pct=row[3].value,
                                                nombre_cp=row[4].value,
                                                pays_brevets=p
                                            )
                                            total_rows += 1
                            # Insertion réussie
                            if total_rows > 0:
                                messages.success(request, "{} brevets ajoutés dans la base de données."
                                                 .format(total_rows), extra_tags='success')
                                return redirect('core:listing_brevets')
                            else:
                                messages.info(request, "Aucune données ajoutées dans la base de données.",
                                              extra_tags='info')
                                return redirect('core:listing_brevets')
                    else:
                        messages.warning(request, "Veuillez choisir un classeur excel comme fichier s'il vous plaît.",
                                         extra_tags='warning')
                        return redirect('core:load_brevets')
                else:
                    messages.warning(request, "Veuillez sélectionner un fichier de Brevets s'il vous plaît.",
                                     extra_tags='warning')
                    return redirect('core:load_brevets')
            else:
                messages.warning(request, "Veuillez entrer une année valide s'il vous plaît.", extra_tags='warning')
                return redirect('core:load_brevets')
        else:
            messages.error(request, "Veuillez remplir tous les champs du formulaire s'il vous plaît.",
                           extra_tags='danger')
            return redirect('core:load_brevets')

    context = {
        'menu': menu,
        'd_menu': d_menu
    }
    return render(request, 'core/load_brevets.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def check_file_dmi(request):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "dmi"
    extensions = ['xls', 'xlsx']
    file_sheets_names = ['Nationaux', 'Etrangers']
    file_name = "DMI"
    error_names = 2
    total_rows = 0

    if request.method == 'POST':
        form = UploadPIFileForm(request.POST or None, request.FILES or None)
        if form.is_valid():
            # On recupere les données du formulaire
            file = form.cleaned_data['file']
            type_pi = form.cleaned_data['type_pi']
            annee = form.cleaned_data['annee']
            # Année valide
            if annee > 1964:
                # nom du fichier correct
                if file_name in str(file).split(".")[0]:
                    # Extensions correcte
                    if str(file).split(".")[1] in extensions:
                        # Lecture du fichier
                        book = load_workbook(file)
                        # On vérifie si les feuilles du classeur sont conforme
                        for n in file_sheets_names:
                            if n not in book.sheetnames:
                                error_names -= 1
                        # Messages d'erreur si les feuilles ne sont pas conforment
                        if error_names == 0:
                            messages.warning(request,
                                             "Certains noms de feuilles du classeur sont incorrectes, veuillez "
                                             "modifier les noms par ceux qui correspondent.", extra_tags='warning')
                            return redirect('core:load_dmi')
                        else:
                            # Insertion de données pour chaque feuille
                            for n in file_sheets_names:
                                for row in book[n]:
                                    if row[1].value is not None:
                                        # Si le pays n'existe pas on le crée
                                        if not Pays.objects.filter(Q(code_pays=row[1].value)).exists():
                                            Pays.objects.create(
                                                code_pays=row[1].value,
                                                nom_pays=row[2].value,
                                                specification_pays=n
                                            )

                                        # on recupère le pays
                                        p = Pays.objects.get(code_pays=row[1].value)

                                        if not DMI.objects.filter(Q(pays_dmi=p) & Q(type_dmi=type_pi)
                                                                  & Q(nombre_dmi=row[3].value)
                                                                  & Q(annee_dmi=annee)).exists():
                                            try:
                                                # On crée le DMI
                                                DMI.objects.create(
                                                    pays_dmi=p,
                                                    type_dmi=type_pi,
                                                    nombre_dmi=row[3].value,
                                                    annee_dmi=annee,
                                                    classe_1=row[10].value,
                                                    classe_2=row[11].value,
                                                    classe_3=row[12].value,
                                                    classe_4=row[13].value,
                                                    classe_5=row[14].value,
                                                    classe_6=row[15].value,
                                                    classe_7=row[16].value,
                                                    classe_8=row[17].value,
                                                    classe_9=row[18].value,
                                                    classe_10=row[19].value,
                                                    classe_11=row[20].value,
                                                    classe_12=row[21].value,
                                                    classe_13=row[22].value,
                                                    classe_14=row[23].value,
                                                    classe_15=row[24].value,
                                                    classe_16=row[25].value,
                                                    classe_17=row[26].value,
                                                    classe_18=row[27].value,
                                                    classe_19=row[28].value,
                                                    classe_20=row[29].value,
                                                    classe_21=row[30].value,
                                                    classe_22=row[31].value,
                                                    classe_23=row[32].value,
                                                    classe_24=row[33].value,
                                                    classe_25=row[34].value,
                                                    classe_26=row[35].value,
                                                    classe_27=row[36].value,
                                                    classe_28=row[37].value,
                                                    classe_29=row[38].value,
                                                    classe_30=row[39].value,
                                                    classe_31=row[40].value,
                                                    classe_32=row[41].value
                                                )
                                            except ValueError:
                                                messages.warning(request, "Veuillez respecter l'écart entre les "
                                                                          "colonnes dans ce fichier s'il vous plaît.",
                                                                 extra_tags='warning')
                                                return redirect('core:load_dmi')
                                            total_rows += 1
                            # Insertion réussie
                            if total_rows > 0:
                                messages.success(request, "{} DMI ajoutés dans la base de données."
                                                 .format(total_rows), extra_tags='success')
                                return redirect('core:listing_dmi')
                            else:
                                messages.info(request, "Aucune données ajoutées dans la base de données.",
                                              extra_tags='info')
                                return redirect('core:listing_dmi')
                    else:
                        messages.warning(request, "Veuillez choisir un classeur excel comme fichier s'il vous plaît.",
                                         extra_tags='warning')
                        return redirect('core:load_dmi')
                else:
                    messages.warning(request, "Veuillez sélectionner un fichier de DMI s'il vous plaît.",
                                     extra_tags='warning')
                    return redirect('core:load_dmi')
            else:
                messages.warning(request, "Veuillez entrer une année valide s'il vous plaît.", extra_tags='warning')
                return redirect('core:load_dmi')
        else:
            messages.error(request, "Veuillez remplir tous les champs du formulaire s'il vous plaît.",
                           extra_tags='danger')
            return redirect('core:load_dmi')

    context = {
        'menu': menu,
        'd_menu': d_menu
    }
    return render(request, 'core/load_dmi.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def check_file_marques(request):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "marque"
    extensions = ['xls', 'xlsx']
    file_sheets_names = {
        'Enregistrement': ['Enreg Madrid', 'Enreg Regionale - Nationale', 'Enreg Regionale - Etranger'],
        'Demande': ['Demande Madrid', 'Demande Regionale - Nationale', 'Demande Regionale - Etranger']
    }
    file_name = "Marque"
    error_names = 2
    total_rows = 0

    if request.method == 'POST':
        form = UploadPIFileForm(request.POST or None, request.FILES or None)
        if form.is_valid():
            # On recupere les données du formulaire
            file = form.cleaned_data['file']
            type_pi = form.cleaned_data['type_pi']
            annee = form.cleaned_data['annee']
            # Année valide
            if annee > 1964:
                # nom du fichier correct
                if file_name in str(file).split(".")[0]:
                    # Extensions correcte
                    if str(file).split(".")[1] in extensions:
                        # Lecture du fichier
                        book = load_workbook(file)
                        # On vérifie si les feuilles du classeur sont conforme
                        for n in file_sheets_names[type_pi]:
                            if n not in book.sheetnames:
                                error_names -= 1
                        # Messages d'erreur si les feuilles ne sont pas conforment
                        if error_names == 0:
                            messages.warning(request,
                                             "Certains noms de feuilles du classeur sont incorrectes, veuillez "
                                             "modifier les noms par ceux qui correspondent.", extra_tags='warning')
                            return redirect('core:load_marques')
                        else:
                            # Insertion de données pour chaque feuille
                            for n in file_sheets_names[type_pi]:
                                try:
                                    for row in book[n]:
                                        if row[1].value is not None:
                                            # Si le pays n'existe pas on le crée
                                            if not Pays.objects.filter(Q(code_pays=row[1].value)).exists():
                                                Pays.objects.create(
                                                    code_pays=row[1].value,
                                                    nom_pays=row[2].value,
                                                )
                                            # on recupère le pays
                                            p = Pays.objects.get(code_pays=row[1].value)

                                            # Si cette marque n'existe pas
                                            if not Marques.objects.filter(Q(pays_marques=p) & Q(type_marques=n)
                                                                          & Q(nombre_marques=row[3].value)
                                                                          & Q(annee_marques=annee)).exists():
                                                try:
                                                    # On crée la marque
                                                    Marques.objects.create(
                                                        pays_marques=p,
                                                        type_marques=n,
                                                        nombre_marques=row[3].value,
                                                        annee_marques=annee,
                                                        classe_1=row[10].value,
                                                        classe_2=row[11].value,
                                                        classe_3=row[12].value,
                                                        classe_4=row[13].value,
                                                        classe_5=row[14].value,
                                                        classe_6=row[15].value,
                                                        classe_7=row[16].value,
                                                        classe_8=row[17].value,
                                                        classe_9=row[18].value,
                                                        classe_10=row[19].value,
                                                        classe_11=row[20].value,
                                                        classe_12=row[21].value,
                                                        classe_13=row[22].value,
                                                        classe_14=row[23].value,
                                                        classe_15=row[24].value,
                                                        classe_16=row[25].value,
                                                        classe_17=row[26].value,
                                                        classe_18=row[27].value,
                                                        classe_19=row[28].value,
                                                        classe_20=row[29].value,
                                                        classe_21=row[30].value,
                                                        classe_22=row[31].value,
                                                        classe_23=row[32].value,
                                                        classe_24=row[33].value,
                                                        classe_25=row[34].value,
                                                        classe_26=row[35].value,
                                                        classe_27=row[36].value,
                                                        classe_28=row[37].value,
                                                        classe_29=row[38].value,
                                                        classe_30=row[39].value,
                                                        classe_31=row[40].value,
                                                        classe_32=row[41].value,
                                                        classe_33=row[42].value,
                                                        classe_34=row[43].value,
                                                        classe_35=row[44].value,
                                                        classe_36=row[45].value,
                                                        classe_37=row[46].value,
                                                        classe_38=row[47].value,
                                                        classe_39=row[48].value,
                                                        classe_40=row[49].value,
                                                        classe_41=row[50].value,
                                                        classe_42=row[51].value,
                                                        classe_43=row[52].value,
                                                        classe_44=row[53].value,
                                                        classe_45=row[54].value
                                                    )
                                                except ValueError:
                                                    messages.warning(request, "Veuillez respecter l'écart entre les "
                                                                              "colonnes dans ce fichier s'il vous plaît.",
                                                                     extra_tags='warning')
                                                    return redirect('core:load_marques')
                                                total_rows += 1
                                except KeyError:
                                    messages.warning(request, "Certains noms de feuilles sont incorects "
                                                              "pour ce fichier.", extra_tags='warning')
                                    return redirect('core:load_marques')
                            # Insertion réussie
                            if total_rows > 0:
                                messages.success(request, "{} Marques ajoutées dans la base de données."
                                                 .format(total_rows), extra_tags='success')
                                return redirect('core:listing_marques')
                            else:
                                messages.info(request, "Aucune données ajoutées dans la base de données.",
                                              extra_tags='info')
                                return redirect('core:listing_marques')
                    else:
                        messages.warning(request, "Veuillez choisir un classeur excel comme fichier s'il vous plaît.",
                                         extra_tags='warning')
                        return redirect('core:load_marques')
                else:
                    messages.warning(request, "Veuillez sélectionner un fichier de Marques s'il vous plaît.",
                                     extra_tags='warning')
                    return redirect('core:load_marques')
            else:
                messages.warning(request, "Veuillez entrer une année valide s'il vous plaît.", extra_tags='warning')
                return redirect('core:load_marques')
        else:
            messages.error(request, "Veuillez remplir tous les champs du formulaire s'il vous plaît.",
                           extra_tags='danger')
            return redirect('core:load_marques')

    context = {
        'menu': menu,
        'd_menu': d_menu
    }
    return render(request, 'core/load_marques.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def load_pays(request):
    # Active menu
    menu = 'pays'

    context = {
        'menu': menu,
    }
    return render(request, 'core/load_pays.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def listing_users(request):
    # Active menu
    menu = 'config'
    i_menu = 'users'

    # On recupère tous les utilisateurs
    users = User.objects.all().order_by('username')

    context = {
        'menu': menu,
        'i_menu': i_menu,
        'users': users,
    }
    return render(request, 'core/listing_users.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def load_brevets(request):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "brevets"

    context = {
        'menu': menu,
        'd_menu': d_menu,
    }
    return render(request, 'core/load_brevets.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def load_dmi(request):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "dmi"

    context = {
        'menu': menu,
        'd_menu': d_menu,
    }
    return render(request, 'core/load_dmi.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def load_marques(request):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "marque"

    context = {
        'menu': menu,
        'd_menu': d_menu,
    }
    return render(request, 'core/load_marques.html', context)


@login_required(login_url='core:login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def listing_pays(request):
    # Infos pour rendre l'option Pays active dans le menu
    menu = "pays"

    # On récupère tous les pays de la base de données
    pays = Pays.objects.all().order_by('nom_pays')
    # Total pays etrangers
    nbre_pays_etrangers = pays.filter(specification_pays='Etrangers').count()
    # Total pays etrangers
    nbre_pays_nationaux = pays.filter(specification_pays='Nationaux').count()

    context = {
        'menu': menu,
        'pays': pays,
        'nbre_pays_etrangers': nbre_pays_etrangers,
        'nbre_pays_nationaux': nbre_pays_nationaux,
    }
    return render(request, 'core/listing_pays.html', context)


@login_required(login_url='core:login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def listing_brevets(request):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "brevets"

    # Tous les brevets triés par ordre croissant de noms de pays
    brevets = Brevets.objects.all().order_by('pays_brevets__nom_pays')
    # Brevets par annnée croissante
    liste_annee_brevets = brevets.values('annee_brevets').order_by('annee_brevets').distinct()
    # Somme PCT et CP
    if len(brevets) > 0:
        nbre_pct = brevets.values('nombre_pct').aggregate(Sum('nombre_pct'))
        nbre_cp = brevets.values('nombre_cp').aggregate(Sum('nombre_cp'))
        # Somme de brevets
        nbre_brevets = nbre_cp['nombre_cp__sum'] + nbre_pct['nombre_pct__sum']
    else:
        nbre_pct = 0
        nbre_cp = 0
        nbre_brevets = 0
    # Nombre de pays ayant des brevets
    nbre_pays = brevets.values('pays_brevets').distinct().aggregate(Count('pays_brevets'))

    # On récupère tous les pays
    pays = Pays.objects.all().order_by('nom_pays')
    load_brevets = []  # Liste de brevets à afficher
    liste_pays_brevets = []  # Liste de pays des brevets
    # On parcourt les Pays afin de charger les données à afficher
    for p in pays:
        # Dictionnaire de données de brevets
        dict_brevets = {}
        # Si le pays a des brevets
        if brevets.filter(pays_brevets=p.pk).exists():
            dict_brevets['pk'] = p.pk
            dict_brevets['code'] = p.code_pays
            dict_brevets['nom'] = p.nom_pays
            dict_brevets['pct'] = p.brevets_set.values('nombre_pct').aggregate(Sum('nombre_pct'))['nombre_pct__sum']
            dict_brevets['cp'] = p.brevets_set.values('nombre_cp').aggregate(Sum('nombre_cp'))['nombre_cp__sum']
            dict_brevets['total'] = p.brevets_set.values('nombre_pct').aggregate(Sum('nombre_pct'))['nombre_pct__sum'] \
                                    + p.brevets_set.values('nombre_cp').aggregate(Sum('nombre_cp'))['nombre_cp__sum']
            liste_pays_brevets.append({'pk': p.pk, 'nom_pays': p.nom_pays})
            load_brevets.append(dict_brevets)  # Ajout des valeurs à la liste
        else:
            continue

    # Données à passer en paramètre à la vue
    context = {
        'menu': menu,
        'd_menu': d_menu,
        'brevets': brevets,
        'liste_annee_brevets': liste_annee_brevets,
        'liste_pays_brevets': liste_pays_brevets,
        'nbre_brevets': nbre_brevets,
        'nbre_pct': nbre_pct,
        'nbre_cp': nbre_cp,
        'nbre_pays': nbre_pays,
        'brevets_list': load_brevets
    }
    return render(request, 'core/listing_brevets.html', context)


@login_required(login_url='core:login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def listing_dmi(request):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "dmi"

    # Tous les DMI de la DB
    dmis = DMI.objects.all().order_by('pays_dmi__nom_pays')
    # Listing des années des DMI
    liste_annee_dmi = dmis.values('annee_dmi').order_by('annee_dmi').distinct()
    # Nombre de pays
    nbre_pays = dmis.values('pays_dmi').distinct().aggregate(Count('pays_dmi'))
    # Nombre de dmi
    nbre_dmi = dmis.values('nombre_dmi').aggregate(Sum('nombre_dmi'))
    # Total des sommes des classes
    total_somme_classes = 0
    for d in dmis:
        total_somme_classes += d.somme_classes
    # Total du nombre de classes
    total_nombre_classes = 0
    for d in dmis:
        total_nombre_classes += d.nombre_classes

    # On récupère tous les pays
    pays = Pays.objects.all().order_by('nom_pays')
    load_dmis = []  # Liste de brevets à afficher
    liste_pays_dmi = []  # Liste de pays des DMI
    # On parcourt les Pays afin de charger les données à afficher
    for p in pays:
        # Dictionnaire de données de brevets
        dict_dmis = {}
        # Si le pays a des brevets
        if dmis.filter(pays_dmi=p.pk).exists():
            dmi = dmis.filter(pays_dmi=p.pk)[0]
            dict_dmis['pk'] = p.pk
            dict_dmis['code'] = p.code_pays
            dict_dmis['nom'] = p.nom_pays
            dict_dmis['somme_classe'] = dmi.somme_classes
            dict_dmis['nombre_classe'] = dmi.nombre_classes
            dict_dmis['total'] = p.dmi_set.values('nombre_dmi')[0]['nombre_dmi']
            liste_pays_dmi.append({'pk': p.pk, 'nom_pays': p.nom_pays})
            load_dmis.append(dict_dmis)  # Ajout des valeurs à la liste
        else:
            continue

    context = {
        'menu': menu,
        'd_menu': d_menu,
        'dmis': dmis,
        'liste_annee_dmi': liste_annee_dmi,
        'liste_pays_dmi': liste_pays_dmi,
        'nbre_pays': nbre_pays,
        'nbre_dmi': nbre_dmi,
        'total_somme_classes': total_somme_classes,
        'total_nombre_classes': total_nombre_classes,
        'dmi_list': load_dmis
    }
    return render(request, 'core/listing_dmi.html', context)


@login_required(login_url='core:login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def listing_marques(request):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "marque"

    # Toutes les marques de la DB
    marques = Marques.objects.all().order_by('pays_marques__nom_pays')
    # Liste années des marques
    liste_annee_marque = marques.values('annee_marques').order_by('annee_marques').distinct()
    # Total de marques
    nbre_marques = marques.values('nombre_marques').aggregate(Sum('nombre_marques'))
    # Nombre de pays
    nbre_pays = marques.values('pays_marques').distinct().aggregate(Count('pays_marques'))
    # Total des classes
    total_somme_classes = 0
    for m in marques:
        total_somme_classes += m.somme_classes
    # Nombre de classes
    total_nombre_classes = 0
    for m in marques:
        total_nombre_classes += m.nombre_classes

    # On récupère tous les pays
    pays = Pays.objects.all().order_by('nom_pays')
    load_marques = []  # Liste de marques à afficher
    liste_pays_marque = []  # Liste de pays des DMI
    # On parcourt les Pays afin de charger les données à afficher
    for p in pays:
        # Dictionnaire de données de brevets
        dict_dmis = {}
        # Si le pays a des brevets
        if marques.filter(pays_marques=p.pk).exists():
            marque = marques.filter(pays_marques=p.pk)[0]
            dict_dmis['pk'] = p.pk
            dict_dmis['code'] = p.code_pays
            dict_dmis['nom'] = p.nom_pays
            dict_dmis['somme_classe'] = marque.somme_classes
            dict_dmis['nombre_classe'] = marque.nombre_classes
            dict_dmis['total'] = p.marques_set.values('nombre_marques')[0]['nombre_marques']
            liste_pays_marque.append({'pk': p.pk, 'nom_pays': p.nom_pays})
            load_marques.append(dict_dmis)  # Ajout des valeurs à la liste
        else:
            continue

    context = {
        'menu': menu,
        'd_menu': d_menu,
        'marques': marques,
        'liste_annee_marque': liste_annee_marque,
        'nbre_marques': nbre_marques,
        'nbre_pays': nbre_pays,
        'total_somme_classes': total_somme_classes,
        'total_nombre_classes': total_nombre_classes,
        'liste_pays_marque': liste_pays_marque,
        'marque_list': load_marques
    }
    return render(request, 'core/listing_marques.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_pays(request):
    # Infos pour rendre l'option Pays active dans le menu
    menu = "pays"

    # Formulaire envoyé
    if request.method == 'POST':
        # On récupère les éléments du formulaire
        form = CountryForm(request.POST)
        # Formulaire valide
        if form.is_valid():
            # Si le pays existe déjà dans la DB
            code_pays = form.cleaned_data['code_pays']
            nom_pays = form.cleaned_data['nom_pays']
            if Pays.objects.filter(Q(code_pays=code_pays) | Q(nom_pays=nom_pays)).exists():
                messages.warning(request, "Le pays que vous essayez d'ajouter existe déjà dans la base de données.",
                                 extra_tags='warning')
                return redirect('core:add_pays')
            else:
                form.save()
                messages.success(request, "Le pays a été ajouté dans la base de données.", extra_tags='success')
                return redirect('core:listing_pays')
        else:
            messages.error(request, "Veuillez remplir tous les champs du formulaire.", extra_tags='danger')
            return redirect('core:add_pays')

    context = {'menu': menu}
    return render(request, 'core/add_pays.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_brevets(request):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "brevets"
    # On récupère les pays de la DB
    pays = Pays.objects.all().order_by('nom_pays')
    # Formulaire validé
    if request.method == 'POST':
        # On recupere les données du formulaire
        form = BrevetsForm(request.POST)
        # Si formulaire bien rempli
        if form.is_valid():
            annee_brevets = form.cleaned_data['annee_brevets']
            pays_pk = form.cleaned_data['pays_brevets']
            type_b = form.cleaned_data['type_brevets']
            # Si le brevet existe déjà dans la DB
            if Brevets.objects.filter(
                    Q(type_brevets=type_b) & Q(annee_brevets=annee_brevets) & Q(pays_brevets=pays_pk)).exists():
                messages.warning(request, "Un brevet de même type existe déjà dans la base de données.",
                                 extra_tags='warning')
                return redirect('core:add_brevets')
            else:
                form.save()
                messages.success(request, "Le brevet a été ajouté dans la base de données.", extra_tags='success')
                return redirect('core:listing_brevets')
        else:
            messages.error(request, "Veuillez remplir tous les champs du formulaire.", extra_tags='danger')
            return redirect('core:add_brevets')

    context = {
        'menu': menu,
        'd_menu': d_menu,
        'pays': pays
    }
    return render(request, 'core/add_brevets.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_dmi(request):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "dmi"

    # On recupere les pays de la DB
    pays = Pays.objects.all().order_by('nom_pays')
    # Formulaire envoyé
    if request.method == 'POST':
        # Recupere le formulaire
        form = DmiForm(request.POST)
        # Si le formulaire est valide
        if form.is_valid():
            # On verifie s'il existe un dmi de meme type
            annee_dmi = form.cleaned_data['annee_dmi']
            pays_pk = form.cleaned_data['pays_dmi']
            type_dmi = form.cleaned_data['type_dmi']
            if DMI.objects.filter(Q(annee_dmi=annee_dmi) & Q(pays_dmi=pays_pk) & Q(type_dmi=type_dmi)).exists():
                messages.warning(request, "Il existe déjà un DMI de même type dans la base de données.",
                                 extra_tags='warning')
                return redirect('core:add_dmi')
            else:
                form.save()
                messages.success(request, "Le DMI a été ajouté dans la base de données.", extra_tags='success')
                return redirect('core:listing_dmi')
        else:
            messages.error(request, "Veuillez remplir tous les champs s'il vous plaît.", extra_tags='danger')
            return redirect('core:add_dmi')
    context = {
        'menu': menu,
        'd_menu': d_menu,
        'pays': pays
    }
    return render(request, 'core/add_dmi.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_marques(request):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "marque"

    # On recupere les pays de la DB
    pays = Pays.objects.all().order_by('nom_pays')
    # Formulaire envoyé
    if request.method == 'POST':
        # Recupere le formulaire
        form = MarqueForm(request.POST)
        # Si le formulaire est valide
        if form.is_valid():
            # On verifie s'il existe un dmi de meme type
            annee_marques = form.cleaned_data['annee_marques']
            pays_pk = form.cleaned_data['pays_marques']
            type_marques = form.cleaned_data['type_marques']
            if Marques.objects.filter(
                    Q(annee_marques=annee_marques) & Q(pays_marques=pays_pk) & Q(type_marques=type_marques)).exists():
                messages.warning(request, "Il existe déjà une Marque de même type dans la base de données.",
                                 extra_tags='warning')
                return redirect('core:add_marques')
            else:
                form.save()
                messages.success(request, "La Marque a été ajoutée dans la base de données.", extra_tags='success')
                return redirect('core:listing_marques')
        else:
            messages.error(request, "Veuillez remplir tous les champs s'il vous plaît.", extra_tags='danger')
            return redirect('core:add_marques')
    context = {
        'menu': menu,
        'd_menu': d_menu,
        'pays': pays
    }
    return render(request, 'core/add_marques.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def edit_status(request, action, username):
    # Active menu
    menu = 'config'
    i_menu = 'users'

    # On recupère l'utilisateur
    try:
        user = User.objects.get(username=username)
    except ObjectDoesNotExist:
        messages.error(request, "L'utilisateur {} n'existe pas dans la base de données.".format(username),
                       extra_tags='danger')
        return redirect('core:register')

    # On effectue l'action passée en paramètre
    if action == 'deactivate':
        user.is_active = False
        user.save()
        messages.info(request, "Le compte de l'utilisateur {} a été désactivé.".format(username),
                      extra_tags='info')
        return redirect('core:register')
    else:
        user.is_active = True
        user.save()
        messages.info(request, "Le compte de l'utilisateur {} a été activé.".format(username),
                      extra_tags='info')
        return redirect('core:register')

    # On recupère tous les utilisateurs
    users = User.objects.all().order_by('username')

    context = {
        'menu': menu,
        'i_menu': i_menu,
        'users': users
    }
    return render(request, 'core/register.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def edit_group(request, old_group, new_group, username):
    # Active menu
    menu = 'config'
    i_menu = 'users'

    # On recupère l'utilisateur
    try:
        user = User.objects.get(username=username)
    except ObjectDoesNotExist:
        messages.error(request, "L'utilisateur {} n'existe pas dans la base de données.".format(username),
                       extra_tags='danger')
        return redirect('core:register')

    # On recupère les groupes passés en paramètres
    try:
        group_old = Group.objects.get(name=old_group)
        group_new = Group.objects.get(name=new_group)
    except ObjectDoesNotExist:
        messages.warning(request, "Veuillez entrer des groupes qui existent au sein de l'administration.",
                         extra_tags='warning')

    # On modifie le groupe
    if new_group == 'admin':
        group_new.user_set.add(user)
        group_old.user_set.remove(user)
        messages.info(request, "L'utilisateur {0} fait désormais parti du groupe {1}.".format(username, new_group),
                      extra_tags='info')
        return redirect('core:register')
    else:
        group_new.user_set.add(user)
        group_old.user_set.remove(user)
        messages.info(request, "L'utilisateur {0} fait désormais parti du groupe {1}.".format(username, new_group),
                      extra_tags='info')
        return redirect('core:register')

    # On recupère tous les utilisateurs
    users = User.objects.all().order_by('username')

    context = {
        'menu': menu,
        'i_menu': i_menu,
        'users': users
    }
    return render(request, 'core/register.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def edit_user(request, username):
    # Active menu
    menu = 'config'

    # On recupère l'utilisateur
    try:
        user = User.objects.get(username=username)
    except ObjectDoesNotExist:
        messages.error(request, "L'utilisateur {} n'existe pas dans la base de données.".format(username),
                       extra_tags='danger')
        return redirect('core:register')

    # Formulaire envoyé
    if request.method == 'POST':
        # Recuperation du formulaire
        form = UpdateUserForm(request.POST, instance=user)
        # Si formulaire valide
        if form.is_valid():
            form.save()
            messages.info(request,
                          "Les informations de {} ont bien été modifiées.".format(form.cleaned_data['username']),
                          extra_tags='info')
            return redirect('core:edit_password', username=form.cleaned_data['username'])
        else:
            messages.error(request, "Veuillez remplir tous les champs.",
                           extra_tags='danger')
            return redirect('core:register')

    context = {
        'menu': menu,
        'user': user
    }
    return render(request, 'core/edit_user.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def edit_password(request, username):
    # Active menu
    menu = 'config'

    # On recupère l'utilisateur
    try:
        user = User.objects.get(username=username)
    except ObjectDoesNotExist:
        messages.error(request, "L'utilisateur {} n'existe pas dans la base de données.".format(username),
                       extra_tags='danger')
        return redirect('core:register')

    # Formulaire envoyé
    if request.method == 'POST':
        # Recuperation du formulaire
        form = UpdatePasswordForm(request.POST)
        # Si formulaire valide
        if form.is_valid():
            # Mot de passe distincts
            if form.cleaned_data['password'] != form.cleaned_data['confirm_password']:
                messages.error(request, "Les mots de passe doivent être identiques.",
                               extra_tags='danger')
                return redirect('core:edit_password', username=username)
            else:
                user.set_password(form.cleaned_data['password'])
                user.save()
                messages.info(request,
                              "Le mot de passe de {} a bien été modifié.".format(username),
                              extra_tags='info')
                return redirect('core:register')
        else:
            messages.error(request, "Veuillez remplir tous les champs.", extra_tags='danger')
            return redirect('core:register')

    context = {
        'menu': menu,
        'user': user
    }
    return render(request, 'core/edit_password.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def edit_pays(request, pk):
    # Infos pour rendre l'option Pays active dans le menu
    menu = "pays"
    # On récupère le pays passé en paramètre
    try:
        pays = Pays.objects.get(pk=pk)
    except ObjectDoesNotExist:
        messages.error(request, "Le pays que vous voulez modifier les données n'existe pas.", extra_tags='danger')
        return redirect('core:listing_pays')

    # Formulaire envoyé
    if request.method == 'POST':
        # Recupère formulaire
        form = CountryForm(request.POST, instance=pays)
        # Formulaire valide
        if form.is_valid():
            pays.nom_pays = form.cleaned_data['nom_pays']
            pays.code_pays = form.cleaned_data['code_pays']
            pays.specification_pays = form.cleaned_data['specification_pays']
            pays.save()
            messages.info(request, "Les données ont été modifiées dans la base de données.", extra_tags='info')
            return redirect('core:listing_pays')
        else:
            messages.error(request, "Impossible d'apporter des modifications à ce pays, vérifiez que le nom du pays "
                                    "comporte moins de 200 caractères", extra_tags='danger')
            return redirect('core:edit_pays', pk=pk)

    context = {
        'pays': pays,
        'menu': menu,
    }
    return render(request, 'core/edit_pays.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def edit_brevets(request, pk):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "brevets"
    # On recupere les pays de la DB
    pays = Pays.objects.all().order_by('nom_pays')
    # On récupère le brevet passé en paramètre
    try:
        brevet = Brevets.objects.get(pk=pk)
    except ObjectDoesNotExist:
        messages.error(request, "Le brevet dont vous voulez modifier les données n'existe pas.", extra_tags='danger')
        return redirect('core:listing_brevets')

    # Formulaire envoyé
    if request.method == 'POST':
        # Recupère formulaire
        form = BrevetsForm(request.POST, instance=brevet)
        # Formulaire valide
        if form.is_valid():
            brevet.type_brevets = form.cleaned_data['type_brevets']
            brevet.annee_brevets = form.cleaned_data['annee_brevets']
            brevet.nombre_pct = form.cleaned_data['nombre_pct']
            brevet.nombre_cp = form.cleaned_data['nombre_cp']
            brevet.pays_brevets = form.cleaned_data['pays_brevets']
            brevet.save()
            messages.info(request, "Les données ont été modifiées dans la base de données.", extra_tags='info')
            return redirect('core:details_brevets', pk=brevet.pays_brevets.pk)
        else:
            messages.error(request,
                           "Impossible d'apporter des modifications à ce brevet, vérifiez que les données que vous "
                           "entrez sont correctes", extra_tags='danger')
            return redirect('core:edit_brevets', pk=brevet.pk)

    context = {
        'menu': menu,
        'd_menu': d_menu,
        'pays': pays,
        'brevet': brevet
    }
    return render(request, 'core/edit_brevets.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def edit_dmi(request, pk):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "dmi"
    # On recupere les pays de la DB
    pays = Pays.objects.all().order_by('nom_pays')
    # On recupere le DMI
    try:
        dmi = DMI.objects.get(pk=pk)
    except ObjectDoesNotExist:
        messages.error(request, "Le DMI dont vous voulez modifier les données n'existe pas.", extra_tags='danger')
        return redirect('core:listing_dmi')

    # Formulaire envoyé
    if request.method == 'POST':
        # On recupere le formulaire
        form = DmiForm(request.POST, instance=dmi)
        # Formulaire valide
        if form.is_valid():
            dmi.pays_dmi = form.cleaned_data['pays_dmi']
            dmi.annee_dmi = form.cleaned_data['annee_dmi']
            dmi.nombre_dmi = form.cleaned_data['nombre_dmi']
            dmi.type_dmi = form.cleaned_data['type_dmi']
            dmi.classe_1 = form.cleaned_data['classe_1']
            dmi.classe_2 = form.cleaned_data['classe_2']
            dmi.classe_3 = form.cleaned_data['classe_3']
            dmi.classe_4 = form.cleaned_data['classe_4']
            dmi.classe_5 = form.cleaned_data['classe_5']
            dmi.classe_6 = form.cleaned_data['classe_6']
            dmi.classe_7 = form.cleaned_data['classe_7']
            dmi.classe_8 = form.cleaned_data['classe_8']
            dmi.classe_9 = form.cleaned_data['classe_9']
            dmi.classe_10 = form.cleaned_data['classe_10']
            dmi.classe_11 = form.cleaned_data['classe_11']
            dmi.classe_12 = form.cleaned_data['classe_12']
            dmi.classe_13 = form.cleaned_data['classe_13']
            dmi.classe_14 = form.cleaned_data['classe_14']
            dmi.classe_15 = form.cleaned_data['classe_15']
            dmi.classe_16 = form.cleaned_data['classe_16']
            dmi.classe_17 = form.cleaned_data['classe_17']
            dmi.classe_18 = form.cleaned_data['classe_18']
            dmi.classe_19 = form.cleaned_data['classe_19']
            dmi.classe_20 = form.cleaned_data['classe_20']
            dmi.classe_21 = form.cleaned_data['classe_21']
            dmi.classe_22 = form.cleaned_data['classe_22']
            dmi.classe_23 = form.cleaned_data['classe_23']
            dmi.classe_24 = form.cleaned_data['classe_24']
            dmi.classe_25 = form.cleaned_data['classe_25']
            dmi.classe_26 = form.cleaned_data['classe_26']
            dmi.classe_27 = form.cleaned_data['classe_27']
            dmi.classe_28 = form.cleaned_data['classe_28']
            dmi.classe_29 = form.cleaned_data['classe_29']
            dmi.classe_30 = form.cleaned_data['classe_30']
            dmi.classe_31 = form.cleaned_data['classe_31']
            dmi.classe_32 = form.cleaned_data['classe_32']
            dmi.save()
            messages.info(request, "Les données ont été modifiées dans la base de données.", extra_tags='info')
            return redirect('core:details_dmi', pk=dmi.pays_dmi.pk)
        else:
            messages.error(request,
                           "Impossible d'apporter des modifications à ce DMI, vérifiez que les données que vous "
                           "entrez sont correctes", extra_tags='danger')
            return redirect('core:edit_brevets', pk=dmi.pk)

    context = {
        'menu': menu,
        'd_menu': d_menu,
        'pays': pays,
        'dmi': dmi,
    }
    return render(request, 'core/edit_dmi.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def edit_marques(request, pk):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "marque"
    # On recupere les pays de la DB
    pays = Pays.objects.all().order_by('nom_pays')
    # On recupere le DMI
    try:
        marque = Marques.objects.get(pk=pk)
    except ObjectDoesNotExist:
        messages.error(request, "La Marque dont vous voulez modifier les données n'existe pas.", extra_tags='danger')
        return redirect('core:listing_dmi')

    # Formulaire envoyé
    if request.method == 'POST':
        # On recupere le formulaire
        form = MarqueForm(request.POST, instance=marque)
        # Formulaire valide
        if form.is_valid():
            marque.pays_marques = form.cleaned_data['pays_marques']
            marque.annee_marques = form.cleaned_data['annee_marques']
            marque.nombre_marques = form.cleaned_data['nombre_marques']
            marque.type_marques = form.cleaned_data['type_marques']
            marque.classe_1 = form.cleaned_data['classe_1']
            marque.classe_2 = form.cleaned_data['classe_2']
            marque.classe_3 = form.cleaned_data['classe_3']
            marque.classe_4 = form.cleaned_data['classe_4']
            marque.classe_5 = form.cleaned_data['classe_5']
            marque.classe_6 = form.cleaned_data['classe_6']
            marque.classe_7 = form.cleaned_data['classe_7']
            marque.classe_8 = form.cleaned_data['classe_8']
            marque.classe_9 = form.cleaned_data['classe_9']
            marque.classe_10 = form.cleaned_data['classe_10']
            marque.classe_11 = form.cleaned_data['classe_11']
            marque.classe_12 = form.cleaned_data['classe_12']
            marque.classe_13 = form.cleaned_data['classe_13']
            marque.classe_14 = form.cleaned_data['classe_14']
            marque.classe_15 = form.cleaned_data['classe_15']
            marque.classe_16 = form.cleaned_data['classe_16']
            marque.classe_17 = form.cleaned_data['classe_17']
            marque.classe_18 = form.cleaned_data['classe_18']
            marque.classe_19 = form.cleaned_data['classe_19']
            marque.classe_20 = form.cleaned_data['classe_20']
            marque.classe_21 = form.cleaned_data['classe_21']
            marque.classe_22 = form.cleaned_data['classe_22']
            marque.classe_23 = form.cleaned_data['classe_23']
            marque.classe_24 = form.cleaned_data['classe_24']
            marque.classe_25 = form.cleaned_data['classe_25']
            marque.classe_26 = form.cleaned_data['classe_26']
            marque.classe_27 = form.cleaned_data['classe_27']
            marque.classe_28 = form.cleaned_data['classe_28']
            marque.classe_29 = form.cleaned_data['classe_29']
            marque.classe_30 = form.cleaned_data['classe_30']
            marque.classe_31 = form.cleaned_data['classe_31']
            marque.classe_32 = form.cleaned_data['classe_32']
            marque.classe_33 = form.cleaned_data['classe_33']
            marque.classe_34 = form.cleaned_data['classe_34']
            marque.classe_35 = form.cleaned_data['classe_35']
            marque.classe_36 = form.cleaned_data['classe_36']
            marque.classe_37 = form.cleaned_data['classe_37']
            marque.classe_38 = form.cleaned_data['classe_38']
            marque.classe_39 = form.cleaned_data['classe_39']
            marque.classe_40 = form.cleaned_data['classe_40']
            marque.classe_41 = form.cleaned_data['classe_41']
            marque.classe_42 = form.cleaned_data['classe_42']
            marque.classe_43 = form.cleaned_data['classe_43']
            marque.classe_44 = form.cleaned_data['classe_44']
            marque.classe_45 = form.cleaned_data['classe_45']
            marque.save()
            messages.info(request, "Les données ont été modifiées dans la base de données.", extra_tags='info')
            return redirect('core:details_marques', pk=marque.pays_marques.pk)
        else:
            messages.error(request,
                           "Impossible d'apporter des modifications à cette Marque, vérifiez que les données que vous "
                           "entrez sont correctes", extra_tags='danger')
            return redirect('core:edit_marques', pk=marque.pk)

    context = {
        'menu': menu,
        'd_menu': d_menu,
        'pays': pays,
        'marque': marque,
    }
    return render(request, 'core/edit_marques.html', context)


@login_required(login_url='core:login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def details_brevets(request, pk):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "brevets"

    # On récupère les brevets du pays
    try:
        brevets = Brevets.objects.filter(pays_brevets=pk).order_by('-annee_brevets')
        pays = Pays.objects.get(pk=pk)
    except ObjectDoesNotExist:
        messages.error(request, "Les brevets dont vous voulez afficher les détails  n'existent pas dans la base "
                                "de données", extra_tags='danger')
        return redirect('core:listing_brevets')

    # Nom du pays
    nom = pays.nom_pays
    # Liste annees de brevets
    liste_annee_brevets = brevets.values('annee_brevets').distinct()
    # Somme PCT
    nbre_pct = brevets.values('nombre_pct').aggregate(Sum('nombre_pct'))
    # Somme CP
    nbre_cp = brevets.values('nombre_cp').aggregate(Sum('nombre_cp'))
    # Somme de brevets
    nbre_brevets = nbre_cp['nombre_cp__sum'] + nbre_pct['nombre_pct__sum']

    load_brevets = []  # Liste de brevets à afficher
    # On parcourt les Pays afin de charger les données à afficher
    for b in brevets:
        # Dictionnaire de données de brevets
        dict_brevets = {
            'pk': b.pk,
            'code': pays.code_pays,
            'pct': b.nombre_pct,
            'cp': b.nombre_cp,
            'annee': b.annee_brevets,
            'type_brevets': b.type_brevets
        }
        load_brevets.append(dict_brevets)

    # Données à passer en paramètre à la vue
    context = {
        'menu': menu,
        'd_menu': d_menu,
        'nom': nom,
        'brevets': brevets,
        'liste_annee_brevets': liste_annee_brevets,
        'nbre_brevets': nbre_brevets,
        'nbre_pct': nbre_pct,
        'nbre_cp': nbre_cp,
        'brevets_list': load_brevets
    }
    return render(request, 'core/details_brevets.html', context)


@login_required(login_url='core:login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def details_dmi(request, pk):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "dmi"

    # On récupère les dmi du pays
    try:
        dmis = DMI.objects.filter(pays_dmi=pk).order_by('-annee_dmi')
        pays = Pays.objects.get(pk=pk)
    except ObjectDoesNotExist:
        messages.error(request, "Les DMI dont vous voulez afficher les détails  n'existent pas dans la base "
                                "de données", extra_tags='danger')
        return redirect('core:listing_dmi')

    # Nom du pays
    nom = pays.nom_pays
    # Liste annees de dmi
    liste_annee_dmi = dmis.values('annee_dmi').distinct().order_by('annee_dmi')
    # Nombre total de DMI
    nbre_dmi = dmis.values('nombre_dmi').aggregate(Sum('nombre_dmi'))
    # Nombre de DMI de type Demande
    nbre_demande = dmis.filter(Q(type_dmi='Demande')).values('nombre_dmi').aggregate(Sum('nombre_dmi'))
    # Nombre de DMI de type Enregistrement
    nbre_enreg = dmis.filter(Q(type_dmi='Enregistrement')).values('nombre_dmi').aggregate(Sum('nombre_dmi'))
    # Total somme de classes
    total_somme_classes = 0
    # Total nombre de classe
    total_nombre_classes = 0
    for d in dmis:
        total_somme_classes += d.somme_classes
        total_nombre_classes += d.nombre_classes

    context = {
        'menu': menu,
        'd_menu': d_menu,
        'nom': nom,
        'dmis': dmis,
        'liste_annee_dmi': liste_annee_dmi,
        'nbre_dmi': nbre_dmi,
        'nbre_demande': nbre_demande,
        'nbre_enreg': nbre_enreg,
        'total_somme_classes': total_somme_classes,
        'total_nombre_classes': total_nombre_classes,
        'dmi_list': dmis
    }
    return render(request, 'core/details_dmi.html', context)


@login_required(login_url='core:login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def details_marques(request, pk):
    # Infos pour rendre l'option PI active dans le menu
    menu = "PI"
    d_menu = "marque"

    # On récupère les marques du pays
    try:
        marques = Marques.objects.filter(pays_marques=pk).order_by('-annee_marques')
        pays = Pays.objects.get(pk=pk)
    except ObjectDoesNotExist:
        messages.error(request, "Les marques dont vous voulez afficher les détails  n'existent pas dans la base "
                                "de données", extra_tags='danger')
        return redirect('core:listing_marques')

    # Nom du pays
    nom = pays.nom_pays
    # Liste annees de dmi
    liste_annee_marques = marques.values('annee_marques').distinct().order_by('annee_marques')
    # Nombre total de DMI
    nbre_marque = marques.values('annee_marques').aggregate(Sum('nombre_marques'))
    # Total somme de classes
    total_somme_classes = 0
    # Total nombre de classe
    total_nombre_classes = 0
    # Total classes services
    total_classes_services = 0
    # Total classes produits
    total_classes_produits = 0
    for d in marques:
        total_somme_classes += d.somme_classes
        total_nombre_classes += d.nombre_classes
        total_classes_produits += d.somme_classes_produits
        total_classes_services += d.somme_classes_services

    # Pagination
    paginator = Paginator(marques, 10)
    page = request.GET.get('page')

    try:
        marque_list = paginator.page(page)
    except PageNotAnInteger:
        marque_list = paginator.page(1)
    except EmptyPage:
        marque_list = paginator.page(paginator.num_pages)

    if len(marques) > 10:
        paginate = True
    else:
        paginate = False

    context = {
        'menu': menu,
        'd_menu': d_menu,
        'nom': nom,
        'marques': marques,
        'liste_annee_marques': liste_annee_marques,
        'nbre_marque': nbre_marque,
        'total_classes_services': total_classes_services,
        'total_classes_produits': total_classes_produits,
        'total_somme_classes': total_somme_classes,
        'total_nombre_classes': total_nombre_classes,
        'marque_list': marque_list,
        'paginate': paginate
    }
    return render(request, 'core/details_marques.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def confirm_delete_pays(request, pk):
    context = {'pk': pk}
    return render(request, 'core/confirm_delete_pays.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def confirm_delete_brevets(request, pk):
    context = {'pk': pk}
    return render(request, 'core/confirm_delete_brevets.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def confirm_delete_dmi(request, pk):
    context = {'pk': pk}
    return render(request, 'core/confirm_delete_dmi.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def confirm_delete_marques(request, pk):
    context = {'pk': pk}
    return render(request, 'core/confirm_delete_marques.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_pays(request, pk):
    if Pays.objects.filter(pk=pk).exists():
        pays = Pays.objects.get(pk=pk)
        pays.delete()
        messages.info(request, "Le pays a été supprimé de la base de données.", extra_tags='info')
    else:
        messages.error(request, "Le pays que vous voulez supprimer n'existe pas dans la base de données",
                       extra_tags='danger')
    return redirect('core:listing_pays')


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_brevets(request, pk):
    if Brevets.objects.filter(pk=pk).exists():
        brevet = Brevets.objects.get(pk=pk)
        brevet.delete()
        messages.info(request, "Le brevet a été supprimé de la base de données.", extra_tags='info')
    else:
        messages.error(request, "Le brevet que vous voulez supprimer n'existe pas dans la base de données",
                       extra_tags='danger')
    return redirect('core:listing_brevets')


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_dmi(request, pk):
    if DMI.objects.filter(pk=pk).exists():
        dmi = DMI.objects.get(pk=pk)
        dmi.delete()
        messages.info(request, "Le DMI a été supprimé de la base de données.", extra_tags='info')
    else:
        messages.error(request, "Le DMI que vous voulez supprimer n'existe pas dans la base de données",
                       extra_tags='danger')
    return redirect('core:listing_dmi')


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_marques(request, pk):
    if Marques.objects.filter(pk=pk).exists():
        marque = Marques.objects.get(pk=pk)
        marque.delete()
        messages.info(request, "La Marque a été supprimé de la base de données.", extra_tags='info')
    else:
        messages.error(request, "La Marque que vous voulez supprimer n'existe pas dans la base de données",
                       extra_tags='danger')
    return redirect('core:listing_marques')


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def suppression_brevets(request):
    if request.method == 'POST':
        form = DeleteBrevetsForm(request.POST)
        if form.is_valid():
            type_pi = form.cleaned_data['type_brevets']
            annee_pi = form.cleaned_data['annee_brevets']
            pi_s = Brevets.objects.filter(Q(type_brevets=type_pi) & Q(annee_brevets=annee_pi))
            if len(pi_s) == 0:
                messages.warning(request, "Il n'existe aucun brevet pour ce type et cette année.",
                                 extra_tags='warning')
                return redirect('core:suppression_brevets')
            else:
                for pi in pi_s:
                    pi.delete()
                messages.success(request, "Les brevets de type {0} de l'année {1} ont été supprimé."
                                 .format(type_pi, annee_pi),
                                 extra_tags='success')
                return redirect('core:listing_brevets')
        else:
            messages.error(request, "Une erreur est survenue lors de la connexion au serveur, impossible de supprimer"
                                    " les données", extra_tags='danger')
            return redirect('core:suppression_brevets')
    else:
        pi_s = Brevets.objects.all()
        liste_type_pi = pi_s.values('type_brevets').distinct()
        liste_annee_pi = pi_s.values('annee_brevets').distinct().order_by('annee_brevets')
        context = {
            'liste_type_pi': liste_type_pi,
            'liste_annee_pi': liste_annee_pi
        }
        return render(request, 'core/suppression_brevets.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def suppression_dmi(request):
    if request.method == 'POST':
        form = DeleteDMIForm(request.POST)
        if form.is_valid():
            type_pi = form.cleaned_data['type_dmi']
            annee_pi = form.cleaned_data['annee_dmi']
            pi_s = DMI.objects.filter(Q(type_dmi=type_pi) & Q(annee_dmi=annee_pi))
            if len(pi_s) == 0:
                messages.warning(request, "Il n'existe aucun DMI pour ce type et cette année.",
                                 extra_tags='warning')
                return redirect('core:suppression_dmi')
            else:
                for pi in pi_s:
                    pi.delete()
                messages.success(request, "Les DMI de type {0} de l'année {1} ont été supprimé."
                                 .format(type_pi, annee_pi),
                                 extra_tags='success')
                return redirect('core:listing_dmi')
        else:
            messages.error(request, "Une erreur est survenue lors de la connexion au serveur, impossible de supprimer"
                                    " les données", extra_tags='danger')
            return redirect('core:suppression_dmi')
    else:
        pi_s = DMI.objects.all()
        liste_type_pi = pi_s.values('type_dmi').distinct()
        liste_annee_pi = pi_s.values('annee_dmi').distinct().order_by('annee_dmi')
        context = {
            'liste_type_pi': liste_type_pi,
            'liste_annee_pi': liste_annee_pi
        }
        return render(request, 'core/suppression_dmi.html', context)


@login_required(login_url='core:login')
@allowed_users(allowed_roles=['admin'])
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def suppression_marques(request):
    if request.method == 'POST':
        form = DeleteMarquesForm(request.POST)
        if form.is_valid():
            type_pi = form.cleaned_data['type_marques']
            annee_pi = form.cleaned_data['annee_marques']
            pi_s = Marques.objects.filter(Q(type_marques=type_pi) & Q(annee_marques=annee_pi))
            if len(pi_s) == 0:
                messages.warning(request, "Il n'existe aucune Marque pour ce type et cette année.",
                                 extra_tags='warning')
                return redirect('core:suppression_marques')
            else:
                for pi in pi_s:
                    pi.delete()
                messages.success(request, "Les marques de type {0} de l'année {1} ont été supprimé."
                                 .format(type_pi, annee_pi),
                                 extra_tags='success')
                return redirect('core:listing_marques')
        else:
            messages.error(request, "Une erreur est survenue lors de la connexion au serveur, impossible de supprimer"
                                    " les données", extra_tags='danger')
            return redirect('core:suppression_marques')
    else:
        pi_s = Marques.objects.all()
        liste_type_pi = pi_s.values('type_marques').distinct()
        liste_annee_pi = pi_s.values('annee_marques').distinct().order_by('annee_marques')
        context = {
            'liste_type_pi': liste_type_pi,
            'liste_annee_pi': liste_annee_pi
        }
        return render(request, 'core/suppression_marques.html', context)
